"""Sổ cái MAIN.xlsx — nguồn chân lý duy nhất của toàn hệ thống.

Ba nguyên tắc bất di bất dịch
-----------------------------
1. **ID không bao giờ đổi.** ID đã được phát cho bộ phận scan và nằm trong tên
   tệp (``ID01.65.1.pdf``). Đổi ID là hỏng toàn bộ dữ liệu đã scan.
2. **Không xóa dòng.** Người rời danh sách chỉ được đánh dấu trạng thái; thư
   mục và ID của họ giữ nguyên.
3. **Không tin dữ liệu dẫn xuất đang lưu.** ``Name_convert`` và ``Folder_name``
   luôn được tính lại từ ``Name``, rồi mới so với giá trị đang lưu. Bằng chứng
   phải làm vậy: dòng ID58 trong MAIN.xlsx gốc lưu ``TrànThịHòngNhuận`` — công
   cụ cũ bóc thiếu dấu.

Ghi tệp
-------
openpyxl ghi đè toàn bộ tệp, đứt giữa chừng là mất sổ cái. Nên luôn ghi ra tệp
tạm rồi ``os.replace()`` (thao tác nguyên tử trên NTFS), kèm xoay vòng 5 bản sao
lưu. Excel khóa tệp đang mở, nên ``PermissionError`` được chuyển thành thông báo
tiếng Việt rõ ràng thay vì để lộ traceback.
"""

from __future__ import annotations

import json
import os
import shutil
import unicodedata
from zipfile import BadZipFile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.paths import thu_muc_du_lieu
from app.core.vietnamese import (
    LoiMaToChuc,
    chuan_hoa_ma_to_chuc,
    con_dau,
    dung_folder_name,
    pascal_case,
    thuoc_dang_bo_co_so,
)

SO_BAN_SAO_LUU = 5

TEN_MAIN_MAC_DINH = "MAIN.xlsx"
# openpyxl chỉ đọc được bốn đuôi này; .xls đời cũ thì không.
DUOI_EXCEL = (".xlsx", ".xlsm", ".xltx", ".xltm")

SHEET_DS = "DSTTHC"
SHEET_CHI_BO = "NAME_FOLDER"
SHEET_DANH_MUC = "DANH_MUC_FILE"

TRANG_THAI_ROI = "Không còn trong danh sách"

# Tiêu đề cột của sheet DSTTHC trong MAIN.xlsx.
# Lưu ý: cột "STT" ở đây là SỐ THỨ TỰ DÒNG, không phải mã loại tài liệu.
# Trong mã nguồn dùng `so_dong` để tránh nhầm với `ma_tai_lieu` (ENG-8).
COT_MAIN = [
    "ID", "STT", "Name", "Name_convert", "Folder_name", "Unit_Folder",
    "DCS_ID", "CCCD_ID", "Ngay_sinh", "Ngay_ket_nap",
    "Chi_bo_dang_sinh_hoat", "Chi_bo_noi_cu_tru", "Trang_thai",
    "Tai_lieu_da_co", "Tai_lieu_chua_co", "Tai_lieu_cho_chuyen_PDF",
    "Tien_do_UT1", "Tien_do_UT2", "Tien_do_UT3",
]

# Tiêu đề bắt buộc phải có trong DS_DANGVIEN.xlsx do hệ thống cấp trên xuất ra.
COT_NGUON = {
    "ho_ten": "Họ và tên",
    "dcs": "Số thẻ Đảng viên",
    "cccd": "Số CCCD",
    "ngay_sinh": "Ngày sinh",
    "ngay_ket_nap": "Ngày kết nạp",
    "chi_bo": "Chi bộ đang sinh hoạt",
    "chi_bo_cu_tru": "Chi bộ nơi cư trú",
    "trang_thai": "Trạng thái",
}


class LoiNghiepVu(Exception):
    """Lỗi có thông báo tiếng Việt, hiển thị thẳng cho người dùng."""


class MainDangBiKhoa(LoiNghiepVu):
    pass


@dataclass
class DongDangVien:
    id: str
    so_dong: int
    name: str
    name_convert: str
    folder_name: str
    unit_folder: str
    dcs_id: str = ""
    cccd_id: str = ""
    ngay_sinh: str = ""
    ngay_ket_nap: str = ""
    chi_bo_dang_sinh_hoat: str = ""
    chi_bo_noi_cu_tru: str = ""
    trang_thai: str = ""
    tai_lieu_da_co: str = ""
    tai_lieu_chua_co: str = ""
    tai_lieu_cho_chuyen_pdf: str = ""
    tien_do_ut1: str = ""
    tien_do_ut2: str = ""
    tien_do_ut3: str = ""

    def thanh_hang(self) -> list:
        return [
            self.id, self.so_dong, self.name, self.name_convert,
            self.folder_name, self.unit_folder, self.dcs_id, self.cccd_id,
            self.ngay_sinh, self.ngay_ket_nap, self.chi_bo_dang_sinh_hoat,
            self.chi_bo_noi_cu_tru, self.trang_thai, self.tai_lieu_da_co,
            self.tai_lieu_chua_co, self.tai_lieu_cho_chuyen_pdf,
            self.tien_do_ut1, self.tien_do_ut2, self.tien_do_ut3,
        ]


@dataclass
class CanhBao:
    muc: str  # "loi" | "canh_bao" | "thong_tin"
    id: str
    ho_ten: str
    van_de: str
    goi_y: str


@dataclass
class SuaDuLieu:
    id: str
    ho_ten: str
    truong: str
    gia_tri_cu: str
    gia_tri_moi: str
    ly_do: str


@dataclass
class KetQuaDongBo:
    dong: list[DongDangVien] = field(default_factory=list)
    them_moi: list[str] = field(default_factory=list)
    roi_danh_sach: list[str] = field(default_factory=list)
    canh_bao: list[CanhBao] = field(default_factory=list)
    du_lieu_ban: list[SuaDuLieu] = field(default_factory=list)

    @property
    def co_loi_chan(self) -> bool:
        return any(c.muc == "loi" for c in self.canh_bao)


# ---------------------------------------------------------------- tiện ích


def _chuoi(v) -> str:
    """Ép giá trị ô Excel về chuỗi, giữ nguyên số 0 đứng đầu của CCCD.

    Bắt buộc chuẩn hóa về NFC
    -------------------------
    Cùng một chuỗi tiếng Việt có hai cách lưu trong Unicode:

        NFC: "Ầ" = U+1EA6                 (một ký tự)
        NFD: "Ầ" = U+0041 U+0302 U+0300   (ba ký tự)

    Hai dạng hiển thị y hệt nhau nhưng ``==`` cho ra False. Dữ liệu thật của
    Viện có lẫn cả hai dạng — đã kiểm chứng: tên ``TRẦN THỊ HỒNG NHUẬN`` đọc
    từ DS_DANGVIEN.xlsx không bằng chính nó gõ từ bàn phím.

    Nếu không chuẩn hóa, phép tra chi bộ ``DS_DANGVIEN.Chi_bo_dang_sinh_hoat``
    ↔ ``NAME_FOLDER.Unit`` có thể trượt và làm cả một chi bộ bị báo "không có
    trong bảng mã", dù tên nhìn giống hệt nhau.

    Chuẩn hóa ngay tại cửa đọc để mọi so sánh phía sau đều an toàn.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return unicodedata.normalize("NFC", str(v)).strip()


def khoa_khop(dcs: str, cccd: str, ho_ten: str, ngay_sinh: str) -> tuple:
    """Khóa dùng để nhận ra "vẫn là người đó" giữa hai lần xuất danh sách.

    Ưu tiên CCCD. Thiếu CCCD thì dùng số thẻ Đảng. Thiếu cả hai (ca ID85 —
    LÊ THỊ THÊM) thì rơi về tên chuẩn hóa cộng ngày sinh.
    """
    if cccd:
        return ("cccd", cccd)
    if dcs:
        return ("dcs", dcs)
    return ("ten_ngaysinh", pascal_case(ho_ten), ngay_sinh)


def _so_tu_id(ma_id: str) -> int:
    so = "".join(c for c in ma_id if c.isdigit())
    return int(so) if so else 0


def id_ke_tiep(da_dung: Iterable[str]) -> str:
    lon_nhat = max((_so_tu_id(i) for i in da_dung), default=0)
    return f"ID{lon_nhat + 1:02d}"


# ----------------------------------------------------------------- đọc vào


TEP_DANH_MUC = thu_muc_du_lieu() / "danh_muc_file.json"

CHU_CAI = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def doc_danh_muc_mac_dinh() -> list[tuple[int, str, str]]:
    """Đọc 104 loại tài liệu từ danh mục đã trích sẵn của app.

    MAIN.xlsx luôn phải có sheet DANH_MUC_FILE. Trước đây sheet này chỉ được ghi
    khi người gọi truyền tham số, nên một lần ghi thiếu tham số đã xóa trắng
    sheet 104 dòng khỏi MAIN.xlsx thật (sự cố 20/8/2026). Giờ danh mục lấy từ
    tệp của app nên không thể quên nữa.
    """
    try:
        du_lieu = json.loads(TEP_DANH_MUC.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return []
    return [
        (int(ma), m["ten_day_du"], "")
        for ma, m in sorted(du_lieu["muc"].items(), key=lambda kv: int(kv[0]))
    ]


def chi_bo_tu_ds(nguon: list[dict[str, str]]) -> list[str]:
    """Rút danh sách chi bộ thực tế đang có trong DS_DANGVIEN, giữ thứ tự xuất hiện.

    Đây mới là danh sách chi bộ đúng — bảng mã chỉ là thứ người dùng điền thêm
    mã tổ chức đảng cho từng chi bộ trong danh sách này.
    """
    thay: list[str] = []
    for r in nguon:
        ten = (r.get("chi_bo") or "").strip()
        if ten and ten not in thay:
            thay.append(ten)
    return thay


def ma_id_theo_thu_tu(so_luong: int) -> list[str]:
    """Sinh mã chữ cái A, B, C… cho từng chi bộ. Quá 26 thì dùng AA, AB…"""
    ra = []
    for i in range(so_luong):
        if i < 26:
            ra.append(CHU_CAI[i])
        else:
            ra.append(CHU_CAI[i // 26 - 1] + CHU_CAI[i % 26])
    return ra


def mo_workbook(duong_dan: Path):
    """Mở tệp Excel, đổi mọi lỗi của openpyxl thành thông báo tiếng Việt.

    openpyxl ném ``InvalidFileException``/``BadZipFile`` kèm thông báo tiếng
    Anh và làm sập cả yêu cầu (HTTP 500 không có thân JSON). Giao diện khi đó
    báo "Không liên lạc được với ứng dụng" — đổ tội cho mạng, trong khi nguyên
    nhân thật chỉ là chọn nhầm tệp, và nó chỉ hiện trong cửa sổ dòng lệnh.
    """
    if duong_dan.is_dir():
        raise LoiNghiepVu(
            f"Đây là thư mục chứ không phải tệp Excel:\n{duong_dan}\n"
            "Chọn đúng tệp .xlsx nằm trong thư mục này."
        )
    if not duong_dan.is_file():
        raise LoiNghiepVu(f"Không tìm thấy tệp:\n{duong_dan}")
    duoi = duong_dan.suffix.lower()
    if duoi not in DUOI_EXCEL:
        mach_nuoc = ""
        if duoi == ".xls":
            mach_nuoc = (
                "\n.xls là định dạng Excel đời cũ. Mở bằng Excel rồi chọn "
                'Save As → "Excel Workbook (*.xlsx)".'
            )
        raise LoiNghiepVu(
            f"Tệp này không phải Excel .xlsx:\n{duong_dan}\n"
            f"Đuôi tệp đang là {duoi or '(không có đuôi)'}.{mach_nuoc}"
        )
    try:
        return load_workbook(duong_dan, read_only=True, data_only=True)
    except PermissionError as loi:
        raise LoiNghiepVu(
            f"Không mở được tệp:\n{duong_dan}\n"
            "Tệp có thể đang mở trong Excel. Đóng Excel rồi thử lại."
        ) from loi
    except (InvalidFileException, BadZipFile) as loi:
        raise LoiNghiepVu(
            f"Tệp hỏng hoặc không phải tệp Excel thật:\n{duong_dan}\n"
            f"Mở thử bằng Excel để kiểm tra.\n(chi tiết: {loi})"
        ) from loi


def chuan_hoa_duong_dan_main(duong_dan: str | Path) -> Path:
    """Chuẩn hóa ô "Nơi lưu sổ cái" người dùng nhập thành đường dẫn tệp .xlsx.

    Người dùng hay đưa vào **thư mục** (đúng nghĩa "nơi lưu") thay vì tệp — app
    tự thêm ``MAIN.xlsx`` chứ không bắt làm lại. Ngược lại, tuyệt đối không nhận
    đuôi khác ``.xlsx``: bước 2 sẽ GHI ĐÈ lên đường dẫn này, trỏ nhầm vào một
    tệp bất kỳ là mất tệp đó.
    """
    tho = str(duong_dan).strip().strip('"')
    if not tho:
        raise LoiNghiepVu("Chưa chọn nơi lưu tệp sổ cái MAIN.xlsx.")
    p = Path(tho)
    if p.is_dir() or tho[-1] in "\\/" or not p.suffix:
        p = p / TEN_MAIN_MAC_DINH
    if p.suffix.lower() != ".xlsx":
        raise LoiNghiepVu(
            f"Sổ cái phải là tệp .xlsx, nhưng đường dẫn đang trỏ tới:\n{p}\n"
            "Bước 2 sẽ ghi đè lên đường dẫn này nên app không nhận đuôi khác.\n"
            "Chọn một thư mục (app tự đặt tên MAIN.xlsx) hoặc gõ đường dẫn kết "
            "thúc bằng .xlsx."
        )
    if p.exists() and not p.is_file():
        raise LoiNghiepVu(f"Đường dẫn sổ cái đang là thư mục chứ không phải tệp:\n{p}")
    return p


def doc_chi_bo(duong_dan: Path) -> dict[str, tuple[str, str]]:
    """Đọc sheet NAME_FOLDER -> {tên chi bộ: (ID chữ cái, mã tổ chức có dấu chấm)}."""
    wb = mo_workbook(duong_dan)
    try:
        if SHEET_CHI_BO not in wb.sheetnames:
            raise LoiNghiepVu(
                f"Tệp {duong_dan.name} không có sheet {SHEET_CHI_BO}. "
                "Đây là bảng mã chi bộ do Đảng ủy UBND tỉnh cung cấp, bắt buộc phải có."
            )
        ra: dict[str, tuple[str, str]] = {}
        for hang, gia_tri in enumerate(wb[SHEET_CHI_BO].iter_rows(values_only=True), 1):
            if hang == 1:
                continue
            ma_id, ma_folder, ten = (_chuoi(v) for v in (list(gia_tri) + ["", "", ""])[:3])
            if not ten:
                continue
            try:
                ra[ten] = (ma_id, chuan_hoa_ma_to_chuc(ma_folder))
            except LoiMaToChuc as loi:
                raise LoiNghiepVu(
                    f'Chi bộ "{ten}" trong sheet {SHEET_CHI_BO}: {loi}'
                ) from loi
        if not ra:
            raise LoiNghiepVu(f"Sheet {SHEET_CHI_BO} trong {duong_dan.name} không có dòng nào.")
        return ra
    finally:
        wb.close()


def doc_ds_dangvien(duong_dan: Path) -> list[dict[str, str]]:
    """Đọc DS_DANGVIEN.xlsx do hệ thống quản lý đảng viên cấp trên xuất ra."""
    wb = mo_workbook(duong_dan)
    try:
        ws = wb[SHEET_DS] if SHEET_DS in wb.sheetnames else wb[wb.sheetnames[0]]
        hang_lap = ws.iter_rows(values_only=True)
        try:
            tieu_de = [_chuoi(v) for v in next(hang_lap)]
        except StopIteration:
            raise LoiNghiepVu(f"Tệp {duong_dan.name} rỗng.") from None

        vi_tri: dict[str, int] = {}
        thieu: list[str] = []
        for khoa, nhan in COT_NGUON.items():
            if nhan in tieu_de:
                vi_tri[khoa] = tieu_de.index(nhan)
            else:
                thieu.append(nhan)
        if thieu:
            raise LoiNghiepVu(
                f"Tệp {duong_dan.name} thiếu cột: {', '.join(thieu)}. "
                "Có thể hệ thống cấp trên đã đổi tên cột — kiểm tra lại bản xuất."
            )

        ra: list[dict[str, str]] = []
        for gia_tri in hang_lap:
            o = [_chuoi(v) for v in gia_tri]
            lay = lambda k: o[vi_tri[k]] if vi_tri[k] < len(o) else ""  # noqa: E731
            if not lay("ho_ten"):
                continue
            ra.append({k: lay(k) for k in COT_NGUON})
        return ra
    finally:
        wb.close()


def chuan_bi_bang_chi_bo(
    nguon: list[dict[str, str]], main_hien_co: Path | None = None
) -> list[dict[str, str]]:
    """Dựng bảng chi bộ để người dùng điền mã tổ chức đảng.

    Theo ``1.Dacta_fixV1``, MAIN.xlsx **được tạo ra từ DS_DANGVIEN**, không phải
    chọn từ một tệp mẫu. Danh sách chi bộ vì thế phải lấy từ chính DS; người
    dùng chỉ điền thêm mã tổ chức đảng cho từng chi bộ.

    Nếu đã có MAIN.xlsx thì điền sẵn mã và **giữ nguyên mã chữ cái cũ** — chữ
    cái đó đã nằm trong tên tệp scan (``A.ID01.65.1``), đổi là hỏng dữ liệu.
    """
    da_co: dict[str, tuple[str, str]] = {}
    if main_hien_co and Path(main_hien_co).is_file():
        try:
            da_co = doc_chi_bo(Path(main_hien_co))
        except LoiNghiepVu:
            da_co = {}

    ten_chi_bo = chi_bo_tu_ds(nguon)
    chu_da_dung = {ma_id for ma_id, _ in da_co.values() if ma_id}
    chu_con_trong = [c for c in CHU_CAI if c not in chu_da_dung]

    ra: list[dict[str, str]] = []
    for ten in ten_chi_bo:
        if ten in da_co:
            ma_id, ma_to_chuc = da_co[ten]
        else:
            ma_id = chu_con_trong.pop(0) if chu_con_trong else ""
            ma_to_chuc = ""
        ra.append(
            {
                "ten": ten,
                "ma_id": ma_id,
                "ma_to_chuc": ma_to_chuc,
                "so_dang_vien": sum(1 for r in nguon if r.get("chi_bo") == ten),
            }
        )
    return ra


def bang_chi_bo_tu_nguoi_dung(
    bang: list[dict], ma_co_so: str = ""
) -> dict[str, tuple[str, str]]:
    """Chuyển bảng người dùng nhập trên giao diện thành dict dùng cho ``dong_bo``.

    Kiểm luôn định dạng mã tổ chức đảng, báo lỗi kèm tên chi bộ để người dùng
    biết sửa dòng nào.

    ``ma_co_so`` là mã đảng bộ cơ sở khai ở bước 0, ví dụ ``38.168.053``. Có mã
    này thì app đối chiếu chéo: ba nhóm số đầu của mã chi bộ phải khớp. Gõ nhầm
    một chữ số ở nhóm giữa sẽ đẩy cả chi bộ sang một cây thư mục khác — lỗi
    lặng lẽ, chỉ lộ ra khi cây đã tạo xong và tệp đã chép vào sai chỗ.
    """
    ra: dict[str, tuple[str, str]] = {}
    thieu: list[str] = []
    lech: list[tuple[str, str]] = []
    mau = f"{ma_co_so}.000.001" if ma_co_so else "38.168.053.000.001"
    for d in bang:
        ten = unicodedata.normalize("NFC", str(d.get("ten", ""))).strip()
        if not ten:
            continue
        ma_to_chuc = str(d.get("ma_to_chuc", "")).strip()
        if not ma_to_chuc:
            thieu.append(ten)
            continue
        try:
            chuan = chuan_hoa_ma_to_chuc(ma_to_chuc)
        except LoiMaToChuc as loi:
            raise LoiNghiepVu(f'Chi bộ "{ten}": {loi}') from loi
        if not thuoc_dang_bo_co_so(chuan, ma_co_so):
            lech.append((ten, chuan))
            continue
        ra[ten] = (str(d.get("ma_id", "")).strip(), chuan)
    if thieu:
        raise LoiNghiepVu(
            "Chưa nhập mã tổ chức đảng cho: "
            + ", ".join(f'"{t}"' for t in thieu)
            + f".\nMã do đảng ủy cấp trên cấp, dạng {mau}"
        )
    if lech:
        chi_tiet = "; ".join(f'"{t}" đang là {m}' for t, m in lech)
        raise LoiNghiepVu(
            f"Mã chi bộ không thuộc đảng bộ cơ sở {ma_co_so} đã khai ở bước 0: "
            f"{chi_tiet}.\nBa nhóm số đầu của mã chi bộ phải đúng bằng "
            f"{ma_co_so} — ví dụ {mau}. Sai một chữ số là cả chi bộ bị đẩy sang "
            f"một cây thư mục khác. Sửa lại bảng, hoặc quay về bước 0 nếu mã "
            f"đảng bộ cơ sở mới là chỗ gõ nhầm."
        )
    trung = [m for m in ra.values()]
    if len({m[1] for m in trung}) != len(trung):
        raise LoiNghiepVu("Có hai chi bộ dùng trùng một mã tổ chức đảng. Kiểm tra lại bảng.")
    return ra


def doc_main(duong_dan: Path) -> list[DongDangVien]:
    """Đọc sổ cái hiện có. Trả về danh sách rỗng nếu tệp chưa tồn tại."""
    if duong_dan.is_dir():
        raise LoiNghiepVu(
            f"Đường dẫn sổ cái đang là thư mục chứ không phải tệp:\n{duong_dan}\n"
            "Chọn tệp MAIN.xlsx, hoặc chọn thư mục rồi để app tự đặt tên MAIN.xlsx."
        )
    if not duong_dan.is_file():
        return []
    wb = mo_workbook(duong_dan)
    try:
        if SHEET_DS not in wb.sheetnames:
            return []
        ra: list[DongDangVien] = []
        for hang, gia_tri in enumerate(wb[SHEET_DS].iter_rows(values_only=True), 1):
            if hang == 1:
                continue
            o = [_chuoi(v) for v in gia_tri] + [""] * len(COT_MAIN)
            if not o[0]:
                continue
            ra.append(
                DongDangVien(
                    id=o[0], so_dong=int(o[1] or hang - 1), name=o[2],
                    name_convert=o[3], folder_name=o[4], unit_folder=o[5],
                    dcs_id=o[6], cccd_id=o[7], ngay_sinh=o[8], ngay_ket_nap=o[9],
                    chi_bo_dang_sinh_hoat=o[10], chi_bo_noi_cu_tru=o[11],
                    trang_thai=o[12], tai_lieu_da_co=o[13], tai_lieu_chua_co=o[14],
                    tai_lieu_cho_chuyen_pdf=o[15], tien_do_ut1=o[16],
                    tien_do_ut2=o[17], tien_do_ut3=o[18],
                )
            )
        return ra
    finally:
        wb.close()


# ------------------------------------------------------------- đồng bộ


def dong_bo(
    nguon: list[dict[str, str]],
    chi_bo: dict[str, tuple[str, str]],
    hien_co: list[DongDangVien],
) -> KetQuaDongBo:
    """Hợp nhất danh sách nguồn vào sổ cái. Không bao giờ đổi hoặc thu hồi ID."""
    kq = KetQuaDongBo()
    theo_khoa = {
        khoa_khop(d.dcs_id, d.cccd_id, d.name, d.ngay_sinh): d for d in hien_co
    }
    da_dung_id = {d.id for d in hien_co}
    con_lai = dict(theo_khoa)

    for so_dong, r in enumerate(nguon, start=1):
        khoa = khoa_khop(r["dcs"], r["cccd"], r["ho_ten"], r["ngay_sinh"])
        cu = con_lai.pop(khoa, None)

        if cu is None:
            ma_id = id_ke_tiep(da_dung_id)
            da_dung_id.add(ma_id)
            kq.them_moi.append(ma_id)
        else:
            ma_id = cu.id

        # 1.Dacta_fixV1: tên thư mục cá nhân dùng SỐ CĂN CƯỚC CÔNG DÂN,
        # không phải số thẻ Đảng. Số thẻ chỉ là phương án dự phòng.
        ma_dinh_danh = r["cccd"] or r["dcs"]
        ten_thu_muc = dung_folder_name(ma_dinh_danh, r["ho_ten"])
        ten_chuan = pascal_case(r["ho_ten"])

        if r["chi_bo"] not in chi_bo:
            kq.canh_bao.append(
                CanhBao(
                    muc="loi", id=ma_id, ho_ten=r["ho_ten"],
                    van_de=f'Chi bộ "{r["chi_bo"]}" không có trong bảng mã NAME_FOLDER.',
                    goi_y="Bổ sung chi bộ này vào sheet NAME_FOLDER, hoặc sửa lại tên chi bộ trong DS_DANGVIEN cho khớp.",
                )
            )
            ma_thu_muc_don_vi = ""
        else:
            ma_thu_muc_don_vi = chi_bo[r["chi_bo"]][1]

        if not ma_dinh_danh:
            kq.canh_bao.append(
                CanhBao(
                    muc="canh_bao", id=ma_id, ho_ten=r["ho_ten"],
                    van_de="Thiếu cả số thẻ Đảng và số CCCD.",
                    goi_y=f'Vẫn tạo thư mục "{ten_thu_muc}" (chỉ có tên). Đề nghị chi bộ bổ sung mã rồi chạy lại.',
                )
            )
        elif not r["cccd"]:
            kq.canh_bao.append(
                CanhBao(
                    muc="thong_tin", id=ma_id, ho_ten=r["ho_ten"],
                    van_de="Chưa có số CCCD trong danh sách nguồn.",
                    goi_y="Đang tạm dùng số thẻ Đảng làm mã thư mục. "
                          "Bổ sung CCCD rồi chạy lại, app sẽ tự đổi tên thư mục.",
                )
            )
        elif len(r["cccd"]) != 12:
            kq.canh_bao.append(
                CanhBao(
                    muc="thong_tin", id=ma_id, ho_ten=r["ho_ten"],
                    van_de=f'Số CCCD "{r["cccd"]}" có {len(r["cccd"])} chữ số, '
                           "quy định là 12 số.",
                    goi_y="Vẫn dùng để đặt tên thư mục. Kiểm tra lại danh sách nguồn.",
                )
            )

        # Không tin dữ liệu dẫn xuất đang lưu: so lại với giá trị vừa tính.
        if cu is not None:
            for truong, cu_gt, moi_gt in (
                ("Name_convert", cu.name_convert, ten_chuan),
                ("Folder_name", cu.folder_name, ten_thu_muc),
            ):
                if cu_gt and cu_gt != moi_gt:
                    ly_do = (
                        f"Giá trị đang lưu còn ký tự có dấu: {' '.join(con_dau(cu_gt))}"
                        if con_dau(cu_gt)
                        else "Giá trị đang lưu khác với giá trị tính lại từ họ tên."
                    )
                    kq.du_lieu_ban.append(
                        SuaDuLieu(ma_id, r["ho_ten"], truong, cu_gt, moi_gt, ly_do)
                    )

        kq.dong.append(
            DongDangVien(
                id=ma_id, so_dong=so_dong, name=r["ho_ten"],
                name_convert=ten_chuan, folder_name=ten_thu_muc,
                unit_folder=ma_thu_muc_don_vi, dcs_id=r["dcs"], cccd_id=r["cccd"],
                ngay_sinh=r["ngay_sinh"], ngay_ket_nap=r["ngay_ket_nap"],
                chi_bo_dang_sinh_hoat=r["chi_bo"], chi_bo_noi_cu_tru=r["chi_bo_cu_tru"],
                trang_thai=r["trang_thai"],
                tai_lieu_da_co=cu.tai_lieu_da_co if cu else "",
                tai_lieu_chua_co=cu.tai_lieu_chua_co if cu else "",
                tai_lieu_cho_chuyen_pdf=cu.tai_lieu_cho_chuyen_pdf if cu else "",
                tien_do_ut1=cu.tien_do_ut1 if cu else "",
                tien_do_ut2=cu.tien_do_ut2 if cu else "",
                tien_do_ut3=cu.tien_do_ut3 if cu else "",
            )
        )

    # Người có trong sổ cái nhưng vắng trong bản xuất mới: giữ lại, chỉ đánh dấu.
    so_dong = len(kq.dong)
    for d in con_lai.values():
        so_dong += 1
        d.so_dong = so_dong
        if d.trang_thai != TRANG_THAI_ROI:
            d.trang_thai = TRANG_THAI_ROI
        kq.roi_danh_sach.append(d.id)
        kq.canh_bao.append(
            CanhBao(
                muc="thong_tin", id=d.id, ho_ten=d.name,
                van_de="Không còn trong bản xuất mới của DS_DANGVIEN.",
                goi_y=f'Giữ nguyên thư mục "{d.folder_name}" và mã {d.id}. Không xóa gì.',
            )
        )
        kq.dong.append(d)

    return kq


# ------------------------------------------------------------------ ghi ra


def _xoay_ban_sao_luu(dich: Path) -> None:
    if not dich.exists():
        return
    for i in range(SO_BAN_SAO_LUU - 1, 0, -1):
        cu = dich.with_suffix(f".bak{i}.xlsx")
        if cu.exists():
            cu.replace(dich.with_suffix(f".bak{i + 1}.xlsx"))
    shutil.copy2(dich, dich.with_suffix(".bak1.xlsx"))


def ghi_main(
    dich: Path,
    dong: list[DongDangVien],
    chi_bo: dict[str, tuple[str, str]],
    danh_muc: list[tuple[int, str, str]] | None = None,
) -> None:
    """Ghi sổ cái ra đĩa theo cách nguyên tử, có sao lưu xoay vòng."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(SHEET_CHI_BO)
    ws.append(["ID", "Ma_Folder", "Unit"])
    for ten, (ma_id, ma_folder) in chi_bo.items():
        ws.append([ma_id, ma_folder, ten])

    ws = wb.create_sheet(SHEET_DS)
    ws.append(COT_MAIN)
    for d in dong:
        ws.append(d.thanh_hang())

    # Sheet danh mục 104 loại tài liệu phải LUÔN có mặt. Bỏ quên tham số này
    # từng làm mất trắng sheet đó khỏi MAIN.xlsx thật (sự cố 20/8/2026).
    ws = wb.create_sheet(SHEET_DANH_MUC)
    ws.append(["Mã loại tài liệu", "Tên tài liệu số hóa", "Ghi chú"])
    for ma, ten, ghi_chu in danh_muc or doc_danh_muc_mac_dinh():
        ws.append([ma, ten, ghi_chu])

    tam = dich.with_suffix(".tmp.xlsx")
    try:
        dich.parent.mkdir(parents=True, exist_ok=True)
        wb.save(tam)
        _xoay_ban_sao_luu(dich)
        os.replace(tam, dich)
    except PermissionError as loi:
        tam.unlink(missing_ok=True)
        raise MainDangBiKhoa(
            f"Không ghi được tệp {dich.name} vì đang bị chương trình khác mở.\n"
            f"Đóng {dich.name} trong Excel rồi bấm Thử lại."
        ) from loi
    finally:
        wb.close()
        tam.unlink(missing_ok=True)
