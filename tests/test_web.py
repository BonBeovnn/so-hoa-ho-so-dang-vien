"""Test đầu-cuối qua tầng web: xác thực, khóa bước, và trọn vẹn bước 1 → 3.

Chạy trên bản sao dữ liệu thật của Viện trong thư mục tạm, không đụng tới tệp gốc.
"""

import shutil
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from app import main as web
from app.core.phien import TEP_CAU_HINH, CauHinh, Phien

GOC = Path(__file__).resolve().parents[2]
DS_THAT = GOC / "Without_APP" / "DS_DANGVIEN.xlsx"
MAIN_THAT = GOC / "With_APP" / "MAIN.xlsx"

can_du_lieu_that = pytest.mark.skipif(
    not DS_THAT.exists(), reason="Không tìm thấy DS_DANGVIEN.xlsx"
)


@pytest.fixture
def phien_sach(tmp_path, monkeypatch):
    """Mỗi test một phiên mới, cấu hình ghi vào thư mục tạm."""
    monkeypatch.setattr(web.PHIEN.__class__, "__init__", Phien.__init__, raising=False)
    moi = Phien(cau_hinh=CauHinh())
    monkeypatch.setattr(web, "PHIEN", moi)
    for mo_dun in ("app.core.phien",):
        __import__(mo_dun)
    monkeypatch.setattr(
        "app.core.phien.TEP_CAU_HINH", tmp_path / "cau_hinh.json", raising=False
    )
    # Vá qua monkeypatch để pytest tự trả lại sau mỗi test. Gán thẳng vào lớp
    # (bản cũ làm thế) thì mọi tệp test chạy sau đều mất khả năng ghi cấu hình,
    # và test nào kiểm việc ghi sẽ xanh giả.
    monkeypatch.setattr(CauHinh, "ghi", lambda self: None)
    yield moi


@pytest.fixture
def khach(phien_sach):
    with TestClient(web.app) as c:
        c.cookies.set(web.TEN_COOKIE, web.TOKEN)
        yield c


@pytest.fixture
def khai_don_vi(khach):
    """Làm xong bước 0 với đúng mã của Viện — mở khóa các bước sau."""
    khach.post(
        "/api/buoc0/luu",
        data={
            "ten_dang_bo": "Đảng bộ Viện Nông nghiệp Thanh Hóa",
            "ten_cap_tren": "Đảng bộ UBND tỉnh Thanh Hóa",
            "ma_tinh": "38",
            "ma_cap_tren": "168",
            "ma_co_so": "053",
            "dia_danh": "Thanh Hóa",
        },
    )
    return khach


def nap_day_du(khach, ds, main):
    """Chay tron ven buoc 1: doc DS -> lay bang chi bo -> doi chieu."""
    bang = khach.post(
        "/api/buoc1/doc_ds",
        data={"duong_dan_ds": str(ds), "duong_dan_main": str(main)},
    ).json()["chi_bo"]
    return khach.post(
        "/api/buoc1/nap",
        json={"duong_dan_ds": str(ds), "duong_dan_main": str(main), "chi_bo": bang},
    ).json()


@pytest.fixture
def nguong_rong(monkeypatch):
    """Nới ngưỡng độ dài đường dẫn cho test đầu-cuối.

    Thư mục tạm của pytest dài tới 88 ký tự, vượt ngưỡng 80 mà app chặn — app
    từ chối là ĐÚNG. Ngưỡng đã được kiểm riêng trong test_paths.py; ở đây ta chỉ
    muốn kiểm luồng nghiệp vụ nên nới ngưỡng ra.
    """
    import app.core.paths as mp

    monkeypatch.setattr(mp, "NGUONG_CANH_BAO_GOC", 200)


@pytest.fixture
def ban_sao(tmp_path):
    """Bản sao DS_DANGVIEN.xlsx và MAIN.xlsx để test ghi mà không hỏng tệp gốc."""
    if not DS_THAT.exists():
        pytest.skip("Không tìm thấy dữ liệu thật")
    ds = tmp_path / "DS_DANGVIEN.xlsx"
    main = tmp_path / "MAIN.xlsx"
    shutil.copy2(DS_THAT, ds)
    shutil.copy2(MAIN_THAT, main)
    return ds, main


class TestXacThuc:
    def test_khong_co_token_thi_bi_tu_choi(self, phien_sach):
        with TestClient(web.app) as c:
            tra = c.get("/buoc/1")
        assert tra.status_code == 403
        assert "Phiên không hợp lệ" in tra.text

    def test_token_sai_cung_bi_tu_choi(self, phien_sach):
        with TestClient(web.app) as c:
            c.cookies.set(web.TEN_COOKIE, "token-gia-mao")
            tra = c.get("/buoc/1")
        assert tra.status_code == 403

    def test_api_cung_doi_token(self, phien_sach):
        """Tab trình duyệt khác trên cùng máy không được gọi API."""
        with TestClient(web.app) as c:
            tra = c.get("/api/duyet?p=C:\\")
        assert tra.status_code == 403

    def test_token_dung_tren_dia_chi_thi_dat_cookie(self, phien_sach):
        with TestClient(web.app) as c:
            tra = c.get(f"/buoc/1?t={web.TOKEN}", follow_redirects=False)
        assert tra.status_code == 303
        assert web.TEN_COOKIE in tra.cookies

    def test_tep_tinh_khong_can_token(self, phien_sach):
        with TestClient(web.app) as c:
            assert c.get("/static/style.css").status_code == 200

    def test_dia_chi_khoi_dong_co_token(self):
        assert web.TOKEN in web.dia_chi_khoi_dong()
        assert web.dia_chi_khoi_dong().startswith("http://127.0.0.1:")

    def test_doc_dung_cong_that_tu_dong_lenh(self, monkeypatch):
        """In sai cong la chan dung nguoi van hanh: dia chi dan vao trinh duyet
        se khong mo duoc gi ma ho khong co cach nao tu doan ra."""
        monkeypatch.setattr(
            web.sys, "argv", ["uvicorn", "app.main:app", "--port", "8877"]
        )
        assert web.cong_dang_chay() == 8877
        assert ":8877/" in web.dia_chi_khoi_dong()

    def test_doc_cong_dang_gan_bang_dau_bang(self, monkeypatch):
        monkeypatch.setattr(web.sys, "argv", ["uvicorn", "--port=9001"])
        assert web.cong_dang_chay() == 9001

    def test_khong_truyen_cong_thi_dung_mac_dinh(self, monkeypatch):
        monkeypatch.setattr(web.sys, "argv", ["uvicorn", "app.main:app"])
        monkeypatch.delenv("UVICORN_PORT", raising=False)
        assert web.cong_dang_chay() == 8000


class TestKhoaBuoc:
    def test_buoc_0_luon_mo(self, khach):
        assert khach.get("/buoc/0").status_code == 200

    def test_buoc_1_bi_khoa_khi_chua_khai_don_vi(self, khach):
        """Chưa biết đơn vị là chưa biết mã gốc cây thư mục, không cho đi tiếp."""
        assert "Chưa mở được bước này" in khach.get("/buoc/1").text

    def test_buoc_1_mo_sau_khi_khai_don_vi(self, khach, khai_don_vi):
        assert "Chưa mở được bước này" not in khach.get("/buoc/1").text

    def test_buoc_3_bi_khoa_khi_chua_lam_buoc_2(self, khach, khai_don_vi):
        tra = khach.get("/buoc/3")
        assert "Chưa mở được bước này" in tra.text

    def test_buoc_khong_ton_tai_thi_ve_buoc_0(self, khach):
        tra = khach.get("/buoc/99", follow_redirects=False)
        assert tra.status_code == 303
        assert tra.headers["location"] == "/buoc/0"

    def test_trang_chu_chuyen_ve_buoc_0(self, khach):
        tra = khach.get("/", follow_redirects=False)
        assert tra.headers["location"] == "/buoc/0"


class TestBuoc0ThongTinDonVi:
    """Bước 0 quyết định tên trên báo cáo và mã gốc của cả cây thư mục."""

    def test_luu_duoc_va_ghep_dung_ma(self, khach):
        d = khach.post(
            "/api/buoc0/luu",
            data={
                "ten_dang_bo": "Đảng bộ Viện Nông nghiệp Thanh Hóa",
                "ten_cap_tren": "Đảng bộ UBND tỉnh Thanh Hóa",
                "ma_tinh": "38",
                "ma_cap_tren": "168",
                "ma_co_so": "53",
                "dia_danh": "Thanh Hóa",
            },
        ).json()
        assert d["ma_dang_bo_co_so"] == "38.168.053"   # tự đệm số 0
        assert d["mau_ma_chi_bo"] == "38.168.053.000.001"

    def test_thieu_ten_dang_bo_thi_bao_loi(self, khach):
        tra = khach.post(
            "/api/buoc0/luu", data={"ten_dang_bo": "   ", "ma_co_so": "053"}
        )
        assert tra.status_code == 400
        assert "Chưa nhập tên đảng bộ" in tra.json()["loi"]

    def test_ma_co_so_qua_dai_thi_bao_loi(self, khach):
        tra = khach.post(
            "/api/buoc0/luu", data={"ten_dang_bo": "Đảng bộ X", "ma_co_so": "0531"}
        )
        assert tra.status_code == 400
        assert "[2].[3].[3]" in tra.json()["loi"]

    def test_khong_khoa_cung_ma_cua_vien(self, khach):
        """Đơn vị ngoài Thanh Hóa phải đổi được cả ba nhóm số."""
        d = khach.post(
            "/api/buoc0/luu",
            data={
                "ten_dang_bo": "Đảng bộ Trung tâm Khuyến nông Hà Tĩnh",
                "ma_tinh": "42",
                "ma_cap_tren": "170",
                "ma_co_so": "011",
            },
        ).json()
        assert d["ma_dang_bo_co_so"] == "42.170.011"

    def test_doi_ma_co_so_thi_dat_lai_cac_buoc_sau(self, khach, phien_sach):
        khach.post(
            "/api/buoc0/luu", data={"ten_dang_bo": "Đảng bộ X", "ma_co_so": "053"}
        )
        phien_sach.cau_hinh.danh_dau_xong(3)
        d = khach.post(
            "/api/buoc0/luu", data={"ten_dang_bo": "Đảng bộ X", "ma_co_so": "099"}
        ).json()
        assert d["da_dat_lai"] is True
        assert phien_sach.cau_hinh.buoc_da_xong == [0]

    def test_giu_nguyen_ma_thi_khong_dat_lai(self, khach, phien_sach):
        khach.post(
            "/api/buoc0/luu", data={"ten_dang_bo": "Đảng bộ X", "ma_co_so": "053"}
        )
        phien_sach.cau_hinh.danh_dau_xong(3)
        d = khach.post(
            "/api/buoc0/luu", data={"ten_dang_bo": "Đảng bộ X đổi tên", "ma_co_so": "053"}
        ).json()
        assert d["da_dat_lai"] is False
        assert 3 in phien_sach.cau_hinh.buoc_da_xong

    def test_ten_dang_bo_hien_tren_dau_trang(self, khach, khai_don_vi):
        assert "Đảng bộ Viện Nông nghiệp Thanh Hóa" in khach.get("/buoc/0").text

    def test_ma_chi_bo_lech_dang_bo_co_so_bi_tu_choi(self, khach, khai_don_vi, ban_sao):
        ds, main = ban_sao
        bang = khach.post(
            "/api/buoc1/doc_ds",
            data={"duong_dan_ds": str(ds), "duong_dan_main": str(main)},
        ).json()["chi_bo"]
        bang[0]["ma_to_chuc"] = "38.168.007.000.001"   # lệch nhóm giữa
        tra = khach.post(
            "/api/buoc1/nap",
            json={"duong_dan_ds": str(ds), "duong_dan_main": str(main), "chi_bo": bang},
        )
        assert tra.status_code == 400
        assert "không thuộc đảng bộ cơ sở" in tra.json()["loi"]


class TestTrangHuongDan:
    def test_mo_duoc_khi_chua_lam_buoc_nao(self, khach):
        tra = khach.get("/huong_dan")
        assert tra.status_code == 200
        assert "Hướng dẫn sử dụng" in tra.text

    def test_co_du_bang_ma_loi(self, khach):
        chu = khach.get("/huong_dan").text
        for ma in ("E01", "E07", "W01", "W02"):
            assert ma in chu

    def test_co_ban_quyen(self, khach):
        assert "© 2026 @anhduc97" in khach.get("/huong_dan").text


class TestDuyetThuMuc:
    def test_khong_truyen_gi_thi_liet_ke_o_dia(self, khach):
        d = khach.get("/api/duyet").json()
        assert d["o_dia"]
        assert d["hien_tai"] is None

    def test_liet_ke_thu_muc_that(self, khach, tmp_path):
        # Khong dat ten "con" — Windows cam vi trung ten thiet bi danh rieng.
        (tmp_path / "thu_muc_con").mkdir()
        d = khach.get("/api/duyet", params={"p": str(tmp_path)}).json()
        assert "thu_muc_con" in d["thu_muc"]

    def test_loc_dung_duoi_tep(self, khach, tmp_path):
        (tmp_path / "a.xlsx").write_text("x", encoding="utf-8")
        (tmp_path / "b.docx").write_text("x", encoding="utf-8")
        d = khach.get("/api/duyet", params={"p": str(tmp_path), "tep": ".xlsx"}).json()
        assert d["tep"] == ["a.xlsx"]

    def test_thu_muc_khong_ton_tai_bao_loi_tieng_viet(self, khach, tmp_path):
        tra = khach.get("/api/duyet", params={"p": str(tmp_path / "khong_co")})
        assert tra.status_code == 400
        assert "Không tìm thấy thư mục" in tra.json()["loi"]


class TestBuoc1:
    def test_tep_khong_ton_tai_bao_loi_ro_rang(self, khach, tmp_path):
        tra = khach.post(
            "/api/buoc1/doc_ds", data={"duong_dan_ds": str(tmp_path / "khong_co.xlsx")}
        )
        assert tra.status_code == 400
        loi = tra.json()["loi"]
        assert "Không tìm thấy tệp danh sách đảng viên" in loi
        assert "Traceback" not in loi

    @can_du_lieu_that
    def test_doc_ds_rut_dung_7_chi_bo_tu_danh_sach_goc(self, khach, ban_sao):
        """1.Dacta_fixV1: MAIN duoc TAO RA tu DS, chi bo lay tu chinh DS."""
        ds, main = ban_sao
        d = khach.post(
            "/api/buoc1/doc_ds",
            data={"duong_dan_ds": str(ds), "duong_dan_main": str(main)},
        ).json()
        assert d["so_dang_vien"] == 85
        assert len(d["chi_bo"]) == 7
        assert sum(c["so_dang_vien"] for c in d["chi_bo"]) == 85
        assert d["main_da_ton_tai"] is True
        # Da co so cai -> ma dien san va CO DAU CHAM
        assert all(c["ma_to_chuc"].count(".") == 4 for c in d["chi_bo"])

    @can_du_lieu_that
    def test_doc_ds_khi_chua_co_so_cai_thi_de_trong_ma(self, khach, ban_sao, tmp_path):
        ds, _ = ban_sao
        d = khach.post(
            "/api/buoc1/doc_ds",
            data={"duong_dan_ds": str(ds), "duong_dan_main": str(tmp_path / "MOI.xlsx")},
        ).json()
        assert d["main_da_ton_tai"] is False
        assert all(c["ma_to_chuc"] == "" for c in d["chi_bo"])
        assert [c["ma_id"] for c in d["chi_bo"]] == list("ABCDEFG")

    @can_du_lieu_that
    def test_thieu_ma_to_chuc_thi_bao_ro_chi_bo_nao(self, khach, ban_sao):
        ds, main = ban_sao
        bang = khach.post(
            "/api/buoc1/doc_ds",
            data={"duong_dan_ds": str(ds), "duong_dan_main": str(main)},
        ).json()["chi_bo"]
        bang[0]["ma_to_chuc"] = ""
        tra = khach.post(
            "/api/buoc1/nap",
            json={"duong_dan_ds": str(ds), "duong_dan_main": str(main), "chi_bo": bang},
        )
        assert tra.status_code == 400
        assert bang[0]["ten"] in tra.json()["loi"]

    @can_du_lieu_that
    def test_ma_to_chuc_sai_dinh_dang_bi_chan(self, khach, ban_sao):
        ds, main = ban_sao
        bang = khach.post(
            "/api/buoc1/doc_ds",
            data={"duong_dan_ds": str(ds), "duong_dan_main": str(main)},
        ).json()["chi_bo"]
        bang[0]["ma_to_chuc"] = "38.168"
        tra = khach.post(
            "/api/buoc1/nap",
            json={"duong_dan_ds": str(ds), "duong_dan_main": str(main), "chi_bo": bang},
        )
        assert tra.status_code == 400
        assert "14 chữ số" in tra.json()["loi"]

    @can_du_lieu_that
    def test_nap_du_lieu_that(self, khach, ban_sao):
        ds, main = ban_sao
        d = nap_day_du(khach, ds, main)
        assert d["so_dang_vien"] == 85
        assert d["so_chi_bo"] == 7
        assert d["them_moi"] == 0
        assert d["roi_danh_sach"] == 0
        assert d["du_lieu_ban"] == 2   # ID58: Name_convert + Folder_name
        assert d["canh_bao"] == 1      # ID85 thiếu cả hai mã
        assert d["co_loi_chan"] is False

    @can_du_lieu_that
    def test_nap_xong_thi_mo_khoa_buoc_2(self, khach, ban_sao):
        ds, main = ban_sao
        nap_day_du(khach, ds, main)
        assert khach.get("/buoc/2").status_code == 200
        assert "Chưa mở được bước này" not in khach.get("/buoc/2").text


class TestBuoc2:
    def test_chua_nap_thi_khong_ghi_duoc(self, khach):
        tra = khach.post("/api/buoc2/ghi")
        assert tra.status_code == 400
        assert "Quay lại bước 1" in tra.json()["loi"]

    @can_du_lieu_that
    def test_bang_hien_thi_dung_ca_ID58_va_ID85(self, khach, ban_sao):
        """Bảng bước 2 phải chỉ ra cả dòng Folder_name hỏng lẫn dòng thiếu mã.

        Kiểm bằng dấu hiệu kỹ thuật chứ không bằng tên người thật: tệp test đi
        lên kho công khai, dữ liệu thật thì không.
        """
        from app.core.mainbook import doc_main
        from app.core.vietnamese import con_dau

        ds, main = ban_sao
        nap_day_du(khach, ds, main)
        html = khach.get("/buoc/2").text

        hong = [x for x in doc_main(main) if con_dau(x.folder_name)]
        assert hong, "MAIN.xlsx thật phải còn ít nhất một Folder_name hỏng"
        for x in hong:
            assert x.folder_name in html, "phải hiện giá trị đang lưu"
        assert "Thiếu cả số thẻ Đảng" in html

    @can_du_lieu_that
    def test_ghi_so_cai_va_sua_duoc_du_lieu_ban(self, khach, ban_sao):
        from app.core.mainbook import doc_main
        from app.core.vietnamese import con_dau

        ds, main = ban_sao
        nap_day_du(khach, ds, main)
        d = khach.post("/api/buoc2/ghi").json()
        assert d["so_dong"] == 85

        sau = doc_main(main)
        assert len(sau) == 85
        assert [x.id for x in sau if con_dau(x.folder_name)] == []
        # Dòng từng hỏng phải được sửa thành dạng [12 số]_[HoTenKhongDau].
        import re

        mau = re.compile(r"(\d{12}_)?[A-Za-z]+")
        con_hong = [x.folder_name for x in sau if not mau.fullmatch(x.folder_name)]
        assert con_hong == [], con_hong

    @can_du_lieu_that
    def test_ghi_xong_van_giu_nguyen_85_ma_ID(self, khach, ban_sao):
        from app.core.mainbook import doc_main

        ds, main = ban_sao
        truoc = {x.cccd_id: x.id for x in doc_main(main) if x.cccd_id}
        nap_day_du(khach, ds, main)
        khach.post("/api/buoc2/ghi")
        for x in doc_main(main):
            if x.cccd_id in truoc:
                assert x.id == truoc[x.cccd_id]

    @can_du_lieu_that
    def test_co_ban_sao_luu_sau_khi_ghi(self, khach, ban_sao):
        ds, main = ban_sao
        nap_day_du(khach, ds, main)
        khach.post("/api/buoc2/ghi")
        assert (main.parent / "MAIN.bak1.xlsx").exists()


class TestBuoc3:
    @can_du_lieu_that
    def test_tron_ven_buoc_1_den_3(self, khach, ban_sao, tmp_path, nguong_rong):
        ds, main = ban_sao
        goc = tmp_path / "HSDV"
        goc.mkdir()

        nap_day_du(khach, ds, main)
        khach.post("/api/buoc2/ghi")

        d = khach.post("/api/buoc3/kiem", data={"duong_dan_goc": str(goc)}).json()
        assert d["don_vi_tao_moi"] == 7
        assert d["co_so_tao_moi"] == 1
        assert d["tom_tat"]["tao_moi"] == 85
        assert d["co_loi"] is False
        assert list(goc.iterdir()) == []  # xem trước không ghi gì

        d2 = khach.post("/api/buoc3/tao").json()
        assert d2["tom_tat"]["tao_moi"] == 85
        # Cay 4 cap: 1 dang bo co so + 7 chi bo + 85 dang vien
        assert [p.name for p in goc.iterdir() if p.is_dir()] == ["38.168.053"]
        assert sum(1 for p in goc.rglob("*") if p.is_dir()) == 93

    @can_du_lieu_that
    def test_chay_lai_buoc_3_khong_tao_them_gi(self, khach, ban_sao, tmp_path, nguong_rong):
        ds, main = ban_sao
        goc = tmp_path / "HSDV"
        goc.mkdir()
        nap_day_du(khach, ds, main)
        khach.post("/api/buoc2/ghi")
        khach.post("/api/buoc3/kiem", data={"duong_dan_goc": str(goc)})
        khach.post("/api/buoc3/tao")

        d = khach.post("/api/buoc3/kiem", data={"duong_dan_goc": str(goc)}).json()
        assert d["tom_tat"]["da_co"] == 85
        assert d["tom_tat"]["tao_moi"] == 0
        assert d["can_thay_doi"] is False

    @can_du_lieu_that
    def test_duong_dan_goc_qua_dai_bi_chan(self, khach, ban_sao, tmp_path):
        from app.core.paths import NGUONG_CANH_BAO_GOC

        ds, main = ban_sao
        nap_day_du(khach, ds, main)
        khach.post("/api/buoc2/ghi")

        sau = tmp_path
        while len(str(sau)) <= NGUONG_CANH_BAO_GOC:
            sau = sau / "thu_muc_long_nhau_dai"
        sau.mkdir(parents=True, exist_ok=True)

        tra = khach.post("/api/buoc3/kiem", data={"duong_dan_goc": str(sau)})
        assert tra.status_code == 400
        assert "gần gốc ổ đĩa hơn" in tra.json()["loi"]

    @can_du_lieu_that
    def test_thu_muc_tam_cua_pytest_bi_chan_o_nguong_mac_dinh(self, khach, ban_sao, tmp_path):
        """Khong noi nguong thi chinh thu muc tam cua pytest cung bi tu choi.

        Day la hanh vi DUNG: duong dan dai lam vuot gioi han 260 ky tu cua Windows.
        """
        from app.core.paths import NGUONG_CANH_BAO_GOC

        ds, main = ban_sao
        goc = tmp_path / "HSDV"
        goc.mkdir()
        if len(str(goc)) <= NGUONG_CANH_BAO_GOC:
            pytest.skip("Thu muc tam ngan hon nguong, khong kiem duoc nhanh nay")
        nap_day_du(khach, ds, main)
        khach.post("/api/buoc2/ghi")
        tra = khach.post("/api/buoc3/kiem", data={"duong_dan_goc": str(goc)})
        assert tra.status_code == 400
        assert "vượt ngưỡng" in tra.json()["loi"]

    def test_chua_xem_truoc_thi_khong_tao_duoc(self, khach):
        tra = khach.post("/api/buoc3/tao")
        assert tra.status_code == 400
        assert "Bấm Xem trước" in tra.json()["loi"]


class TestChonNhamNoiLuuSoCai:
    """Bước 1 phải sống sót khi ô "nơi lưu sổ cái" bị điền sai.

    Lỗi thật ngày 20/8/2026: điền một THƯ MỤC vào ô này làm sập HTTP 500 và
    giao diện báo nhầm là mất liên lạc với ứng dụng.
    """

    @can_du_lieu_that
    def test_dien_thu_muc_thi_app_tu_them_ten_tep(self, khach, ban_sao, tmp_path):
        ds, _ = ban_sao
        kho = tmp_path / "kho_so_cai"
        kho.mkdir()
        bang = khach.post(
            "/api/buoc1/doc_ds", data={"duong_dan_ds": str(ds), "duong_dan_main": str(kho)}
        ).json()["chi_bo"]
        # Thu muc trong nen chua co so cai cu: tu dien ma to chuc cho tung chi bo.
        for i, cb in enumerate(bang, 1):
            cb["ma_to_chuc"] = f"38.168.053.000.{i:03d}"
        tra = khach.post(
            "/api/buoc1/nap",
            json={"duong_dan_ds": str(ds), "duong_dan_main": str(kho), "chi_bo": bang},
        )
        assert tra.status_code == 200
        assert tra.json()["duong_dan_main"] == str(kho / "MAIN.xlsx")

    @can_du_lieu_that
    def test_dien_tep_khac_thi_bao_loi_400_chu_khong_sap_500(self, khach, ban_sao, tmp_path):
        ds, _ = ban_sao
        dac_ta = tmp_path / "1.Dacta_fixV1"
        goc = "dac ta cua nguoi dung, khong duoc dung toi"
        dac_ta.write_text(goc, encoding="utf-8")
        bang = khach.post(
            "/api/buoc1/doc_ds", data={"duong_dan_ds": str(ds), "duong_dan_main": ""}
        ).json()["chi_bo"]
        tra = khach.post(
            "/api/buoc1/nap",
            json={"duong_dan_ds": str(ds), "duong_dan_main": str(dac_ta), "chi_bo": bang},
        )
        assert tra.status_code == 400
        assert "phải là tệp .xlsx" in tra.json()["loi"]
        assert dac_ta.read_text(encoding="utf-8") == goc

    @can_du_lieu_that
    def test_doc_ds_van_chay_du_o_noi_luu_gõ_sai(self, khach, ban_sao, tmp_path):
        """Ô "nơi lưu" mới là gợi ý ở bước đọc DS, không được chặn việc đọc."""
        ds, _ = ban_sao
        tra = khach.post(
            "/api/buoc1/doc_ds",
            data={"duong_dan_ds": str(ds), "duong_dan_main": str(tmp_path / "ghi_chu.txt")},
        )
        assert tra.status_code == 200
        assert tra.json()["duong_dan_main"] == str(ds.parent / "MAIN.xlsx")


class TestLoiNgoaiDuTinh:
    def test_van_tra_ve_json_de_giao_dien_khong_do_toi_cho_mang(self, phien_sach, monkeypatch):
        """Không có bộ bắt lỗi này thì 500 rỗng bị giao diện dịch thành
        "Không liên lạc được với ứng dụng" — sai hướng dò lỗi hoàn toàn."""

        def no_tung(*a, **k):
            raise RuntimeError("hong bat ngo")

        monkeypatch.setattr(web, "doc_ds_dangvien", no_tung)
        with TestClient(web.app, raise_server_exceptions=False) as c:
            c.cookies.set(web.TEN_COOKIE, web.TOKEN)
            tra = c.post(
                "/api/buoc1/doc_ds",
                data={"duong_dan_ds": str(DS_THAT), "duong_dan_main": ""},
            )
        assert tra.status_code == 500
        than = tra.json()
        assert "RuntimeError" in than["loi"]
        assert "cửa sổ dòng lệnh" in than["loi"]


@pytest.fixture
def da_tao_cay(khach, ban_sao, tmp_path, nguong_rong):
    """Chạy trọn bước 1→3 rồi trả về (thư mục scan rỗng, thư mục kho, ID thật)."""
    ds, main = ban_sao
    goc = tmp_path / "HSDV"
    goc.mkdir()
    scan = tmp_path / "Scan"
    scan.mkdir()

    nap_day_du(khach, ds, main)
    khach.post("/api/buoc2/ghi")
    khach.post("/api/buoc3/kiem", data={"duong_dan_goc": str(goc)})
    khach.post("/api/buoc3/tao")
    return scan, goc


class TestBuoc4Quet:
    @can_du_lieu_that
    def test_quet_phan_loai_dung_hop_le_va_loi(self, khach, da_tao_cay):
        scan, goc = da_tao_cay
        (scan / "ID01.65.1.pdf").write_bytes(b"tep dung")
        (scan / "Hồ sơ đảng viên Lư Xuân Bắc.doc").write_bytes(b"tep sai ten")

        tra = khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
        assert tra.status_code == 200
        d = tra.json()
        assert (d["tong"], d["so_hop_le"], d["so_loi"]) == (2, 1, 1)
        assert d["dong_loi"][0]["ma_loi"] == "E01"
        assert len(d["dang_vien"]) == 85
        assert len(d["tai_lieu"]) == 104

    @can_du_lieu_that
    def test_khong_cho_quet_thu_muc_nam_trong_kho(self, khach, da_tao_cay):
        """Chép tệp lên chính nó là hỏng dữ liệu, phải chặn từ đầu."""
        scan, goc = da_tao_cay
        ben_trong = goc / "scan_tam"
        ben_trong.mkdir()
        tra = khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(ben_trong)})
        assert tra.status_code == 400
        assert "lồng vào nhau" in tra.json()["loi"]

    @can_du_lieu_that
    def test_tro_vao_tep_thi_bao_ro(self, khach, da_tao_cay):
        scan, goc = da_tao_cay
        tep = scan / "ID01.65.1.pdf"
        tep.write_bytes(b"x")
        tra = khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(tep)})
        assert tra.status_code == 400
        assert "không phải thư mục" in tra.json()["loi"]

    @can_du_lieu_that
    def test_sua_tai_cho_lam_giam_so_loi(self, khach, da_tao_cay):
        scan, goc = da_tao_cay
        sai = scan / "Hồ sơ đảng viên.doc"
        sai.write_bytes(b"tep sai ten")
        khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})

        tra = khach.post(
            "/api/buoc4/sua",
            json={"duong_dan": str(sai), "id_dang_vien": "ID01", "ma_tai_lieu": 87},
        )
        assert tra.status_code == 200
        d = tra.json()
        assert d["so_loi_con_lai"] == 0
        assert d["tep"]["id_dang_vien"] == "ID01"
        assert sai.exists()      # tệp gốc không bị đổi tên

    @can_du_lieu_that
    def test_sua_sang_ma_tai_lieu_la_thi_bi_tu_choi(self, khach, da_tao_cay):
        scan, goc = da_tao_cay
        sai = scan / "abc.pdf"
        sai.write_bytes(b"x")
        khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
        tra = khach.post(
            "/api/buoc4/sua",
            json={"duong_dan": str(sai), "id_dang_vien": "ID01", "ma_tai_lieu": 999},
        )
        assert tra.status_code == 400

    @can_du_lieu_that
    def test_xuat_bang_loi_ra_xlsx(self, khach, da_tao_cay):
        scan, goc = da_tao_cay
        (scan / "linh tinh.pdf").write_bytes(b"x")
        khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
        tra = khach.get("/api/buoc4/xuat_loi")
        assert tra.status_code == 200
        assert tra.content[:2] == b"PK"          # tệp .xlsx là một gói zip
        assert "attachment" in tra.headers["content-disposition"]


class TestBuoc5LuanChuyen:
    @can_du_lieu_that
    def test_tron_ven_xem_truoc_roi_thuc_thi(self, khach, da_tao_cay, tmp_path, monkeypatch):
        from app.core import intake as mo_dun_intake

        monkeypatch.setattr(mo_dun_intake, "THU_MUC_MANIFEST", tmp_path / "nhat_ky")
        scan, goc = da_tao_cay
        (scan / "ID01.2.1.pdf").write_bytes(b"ly lich dang vien")
        (scan / "ID01.1.1.jpg").write_bytes(b"anh chup ly lich")
        khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})

        d = khach.post("/api/buoc5/xem").json()
        assert d["tom_tat"]["copy"] == 2
        assert d["so_dong"] == 2
        assert sum(1 for p in goc.rglob("*") if p.is_file()) == 0   # xem trước không ghi gì

        d2 = khach.post("/api/buoc5/thuc_thi").json()
        assert d2["tom_tat"]["copy"] == 2
        assert Path(d2["manifest"]).is_file()

        tep_dich = sorted(p.name for p in goc.rglob("*") if p.is_file())
        assert tep_dich == [
            "001.Ly_lich_nguoi_xin_vao_Dang.1.jpg",
            "002.Ly_lich_dang_vien.1.pdf",
        ]
        assert any(p.name == "_CHO_CHUYEN_PDF" for p in goc.rglob("*") if p.is_dir())

    @can_du_lieu_that
    def test_chua_xem_truoc_thi_khong_thuc_thi_duoc(self, khach, da_tao_cay):
        scan, goc = da_tao_cay
        (scan / "ID01.2.1.pdf").write_bytes(b"x")
        khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
        tra = khach.post("/api/buoc5/thuc_thi")
        assert tra.status_code == 400
        assert "Xem trước" in tra.json()["loi"]

    @can_du_lieu_that
    def test_chay_lai_lan_hai_khong_nhan_ban(self, khach, da_tao_cay, tmp_path, monkeypatch):
        from app.core import intake as mo_dun_intake

        monkeypatch.setattr(mo_dun_intake, "THU_MUC_MANIFEST", tmp_path / "nhat_ky")
        scan, goc = da_tao_cay
        (scan / "ID01.2.1.pdf").write_bytes(b"ly lich dang vien")
        for _ in range(2):
            khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
            khach.post("/api/buoc5/xem")
            khach.post("/api/buoc5/thuc_thi")
        assert sum(1 for p in goc.rglob("*") if p.is_file()) == 1

    @can_du_lieu_that
    def test_mo_lai_app_sau_khi_tat_thi_bat_quet_lai(self, khach, da_tao_cay, phien_sach):
        """Kết quả quét chỉ nằm trong bộ nhớ. Tắt app là phải quét lại từ bước 4.

        Cấu hình "bước 4 đã xong" thì ghi ra đĩa, nên nếu không chuyển hướng,
        trang bước 5 sẽ hiện ra với bảng rỗng và nút Thực thi vô nghĩa.
        """
        scan, goc = da_tao_cay
        (scan / "ID01.2.1.pdf").write_bytes(b"x")
        khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
        assert khach.get("/buoc/5", follow_redirects=False).status_code == 200

        phien_sach.ket_qua_quet = None      # như vừa tắt rồi mở lại app
        tra = khach.get("/buoc/5", follow_redirects=False)
        assert tra.status_code == 303
        assert tra.headers["location"].endswith("/buoc/4")


@pytest.fixture
def da_luan_chuyen(khach, da_tao_cay, tmp_path, monkeypatch):
    """Chạy hết bước 1→5 với vài tệp thật, sẵn sàng cho bước đối soát."""
    from app.core import intake as mo_dun_intake

    monkeypatch.setattr(mo_dun_intake, "THU_MUC_MANIFEST", tmp_path / "nhat_ky")
    scan, goc = da_tao_cay
    (scan / "ID01.1.1.pdf").write_bytes(b"ly lich nguoi xin vao Dang")
    (scan / "ID01.2.1.pdf").write_bytes(b"ly lich dang vien")
    (scan / "ID02.2.1.pdf").write_bytes(b"ly lich dang vien cua nguoi khac")
    (scan / "ID01.87.1.docx").write_bytes(b"quyet dinh dieu dong")
    khach.post("/api/buoc4/quet", data={"duong_dan_scan": str(scan)})
    khach.post("/api/buoc5/xem")
    khach.post("/api/buoc5/thuc_thi")
    return scan, goc


class TestBuoc6DoiSoat:
    @can_du_lieu_that
    def test_doi_soat_dem_dung_tep_da_chep(self, khach, da_luan_chuyen):
        tra = khach.post("/api/buoc6/doi_soat")
        assert tra.status_code == 200
        d = tra.json()
        assert d["so_dang_vien"] == 85
        assert d["tong_tep_da_co"] == 3          # 3 tệp PDF, tệp .docx không tính
        assert d["so_cho_chuyen_pdf"] == 1
        assert d["tien_do"]["1"]["co"] == 3
        assert d["tien_do"]["1"]["tong"] == 85 * 36

    @can_du_lieu_that
    def test_tung_dang_vien_co_tien_do_rieng(self, khach, da_luan_chuyen):
        d = khach.post("/api/buoc6/doi_soat").json()
        theo_id = {x["id"]: x for x in d["dong"]}
        assert theo_id["ID01"]["ut1"] == "2/36"
        assert theo_id["ID02"]["ut1"] == "1/36"
        assert theo_id["ID03"]["ut1"] == "0/36"
        assert theo_id["ID01"]["cho_chuyen_pdf"] == 1

    @can_du_lieu_that
    def test_ghi_sau_cot_doi_soat_vao_so_cai(self, khach, da_luan_chuyen, ban_sao):
        from openpyxl import load_workbook

        ds, main = ban_sao
        khach.post("/api/buoc6/doi_soat")
        tra = khach.post("/api/buoc6/ghi")
        assert tra.status_code == 200
        assert tra.json()["so_dong"] == 85

        wb = load_workbook(main, read_only=True, data_only=True)
        ws = wb["DSTTHC"]
        tieu_de = [o for o in next(ws.iter_rows(values_only=True))]
        hang = {h[0]: h for h in ws.iter_rows(min_row=2, values_only=True)}
        cot = {ten: i for i, ten in enumerate(tieu_de)}
        assert hang["ID01"][cot["Tai_lieu_da_co"]] == "1,2"
        assert hang["ID01"][cot["Tien_do_UT1"]] == "2/36"
        assert hang["ID01"][cot["Tai_lieu_cho_chuyen_PDF"]] == "87"
        assert str(hang["ID01"][cot["Tai_lieu_chua_co"]]).startswith("3,4,5")
        wb.close()

    @can_du_lieu_that
    def test_chua_doi_soat_thi_khong_ghi_duoc(self, khach, da_luan_chuyen):
        tra = khach.post("/api/buoc6/ghi")
        assert tra.status_code == 400
        assert "Đối soát" in tra.json()["loi"]


class TestBuoc7BaoCao:
    @can_du_lieu_that
    def test_xuat_du_hai_tep_va_tai_ve_duoc(self, khach, da_luan_chuyen, tmp_path):
        scan, goc = da_luan_chuyen
        khach.post("/api/buoc6/doi_soat")
        ra = tmp_path / "bao_cao"
        ra.mkdir()

        tra = khach.post(
            "/api/buoc7/xuat",
            json={"thu_muc_ra": str(ra), "ngay": "2026-08-20", "ho_ten_ky": "Nguyễn Văn A"},
        )
        assert tra.status_code == 200
        d = tra.json()
        assert (ra / d["docx"]).is_file() and (ra / d["xlsx"]).is_file()
        assert d["so_dang_vien"] == 85

        tai = khach.get("/api/buoc7/tai?loai=docx")
        assert tai.status_code == 200
        assert tai.content[:2] == b"PK"

    @can_du_lieu_that
    def test_xuat_rieng_mot_chi_bo(self, khach, da_luan_chuyen, tmp_path):
        khach.post("/api/buoc6/doi_soat")
        d0 = khach.post("/api/buoc6/doi_soat").json()
        ma = d0["chi_bo"][0]["ma_to_chuc"]
        so_dv = d0["chi_bo"][0]["so_dang_vien"]

        ra = tmp_path / "bao_cao_chi_bo"
        ra.mkdir()
        d = khach.post(
            "/api/buoc7/xuat", json={"thu_muc_ra": str(ra), "pham_vi": ma}
        ).json()
        assert d["so_dang_vien"] == so_dv
        assert d["ten_pham_vi"] == d0["chi_bo"][0]["ten"]

    @can_du_lieu_that
    def test_ma_chi_bo_la_thi_bao_loi(self, khach, da_luan_chuyen, tmp_path):
        khach.post("/api/buoc6/doi_soat")
        tra = khach.post(
            "/api/buoc7/xuat",
            json={"thu_muc_ra": str(tmp_path), "pham_vi": "38.168.053.000.999"},
        )
        assert tra.status_code == 400

    @can_du_lieu_that
    def test_ngay_thang_go_sai_thi_noi_ro_dang_dung(self, khach, da_luan_chuyen, tmp_path):
        khach.post("/api/buoc6/doi_soat")
        tra = khach.post(
            "/api/buoc7/xuat", json={"thu_muc_ra": str(tmp_path), "ngay": "20/08/2026"}
        )
        assert tra.status_code == 400
        assert "2026-08-20" in tra.json()["loi"]

    @can_du_lieu_that
    def test_chua_xuat_thi_khong_tai_duoc(self, khach, da_luan_chuyen):
        tra = khach.get("/api/buoc7/tai?loai=docx")
        assert tra.status_code == 400

    @can_du_lieu_that
    def test_mo_lai_app_thi_bat_doi_soat_lai(self, khach, da_luan_chuyen, phien_sach):
        khach.post("/api/buoc6/doi_soat")
        khach.post("/api/buoc6/ghi")
        assert khach.get("/buoc/7", follow_redirects=False).status_code == 200

        phien_sach.ket_qua_doi_soat = None
        tra = khach.get("/buoc/7", follow_redirects=False)
        assert tra.status_code == 303
        assert tra.headers["location"].endswith("/buoc/6")


class TestGoiChanDoan:
    def test_nut_chan_doan_co_o_moi_trang(self, khach):
        html = khach.get("/buoc/1").text
        assert "/api/chan_doan" in html

    def test_tai_ve_duoc_tep_zip(self, khach):
        tra = khach.get("/api/chan_doan")
        assert tra.status_code == 200
        assert tra.content[:2] == b"PK"
        assert "GoiChanDoan_" in tra.headers["content-disposition"]

    @can_du_lieu_that
    def test_goi_ghi_lai_trang_thai_that_cua_phien(self, khach, da_luan_chuyen):
        import io
        import zipfile

        khach.post("/api/buoc6/doi_soat")
        tra = khach.get("/api/chan_doan")
        goi = zipfile.ZipFile(io.BytesIO(tra.content))
        chu = goi.read("trang_thai.txt").decode("utf-8")
        assert "85 đảng viên" in chu
        assert "Bước 4:" in chu and "Bước 6:" in chu

    @can_du_lieu_that
    def test_goi_khong_mang_theo_cccd_that(self, khach, da_luan_chuyen, ban_sao):
        """Chạy trên dữ liệu thật của Viện rồi dò từng số CCCD trong gói."""
        import io
        import zipfile

        from app.core.mainbook import doc_main

        ds, main = ban_sao
        khach.post("/api/buoc6/doi_soat")
        cccd = [d.cccd_id for d in doc_main(main) if d.cccd_id]
        assert len(cccd) > 50          # phải có dữ liệu thật thì phép thử mới có nghĩa

        goi = zipfile.ZipFile(io.BytesIO(khach.get("/api/chan_doan").content))
        toan_bo = b"".join(goi.read(t) for t in goi.namelist())
        for so in cccd:
            assert so.encode() not in toan_bo

