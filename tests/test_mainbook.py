"""Test sổ cái MAIN.xlsx.

Chia hai nhóm:
  * Nhóm logic — chạy trên dữ liệu dựng sẵn, luôn chạy được.
  * Nhóm dữ liệu thật — đọc thẳng DS_DANGVIEN.xlsx và MAIN.xlsx của Viện,
    tự bỏ qua nếu không tìm thấy tệp.
"""

import os
import unicodedata
from pathlib import Path

import pytest

from app.core import mainbook as mb
from app.core.mainbook import (
    DongDangVien,
    LoiNghiepVu,
    MainDangBiKhoa,
    doc_chi_bo,
    doc_ds_dangvien,
    doc_main,
    dong_bo,
    ghi_main,
    id_ke_tiep,
    khoa_khop,
)

GOC = Path(__file__).resolve().parents[2]
DS_THAT = GOC / "Without_APP" / "DS_DANGVIEN.xlsx"
MAIN_THAT = GOC / "With_APP" / "MAIN.xlsx"

# 1.Dacta_fixV1: ma to chuc dang PHAI co dau cham (Quy dinh 208-QD/TW).
CHI_BO = {
    "Chi bộ Văn phòng": ("A", "38.168.053.000.001"),
    "Chi bộ Phòng quản lý khoa học": ("B", "38.168.053.000.002"),
}


def nguoi(ho_ten, dcs="", cccd="", chi_bo="Chi bộ Văn phòng", ngay_sinh="01/01/1990"):
    return {
        "ho_ten": ho_ten, "dcs": dcs, "cccd": cccd, "ngay_sinh": ngay_sinh,
        "ngay_ket_nap": "01/01/2020", "chi_bo": chi_bo,
        "chi_bo_cu_tru": "", "trang_thai": "Đang sinh hoạt",
    }


class TestKhoaKhop:
    def test_uu_tien_cccd(self):
        assert khoa_khop("111", "222", "A B", "01/01/1990") == ("cccd", "222")

    def test_thieu_cccd_thi_dung_so_the_dang(self):
        assert khoa_khop("111", "", "A B", "01/01/1990") == ("dcs", "111")

    def test_thieu_ca_hai_thi_dung_ten_va_ngay_sinh(self):
        """Ca ID85 — LÊ THỊ THÊM không có mã nào."""
        assert khoa_khop("", "", "LÊ THỊ THÊM", "31/05/1969") == (
            "ten_ngaysinh", "LeThiThem", "31/05/1969",
        )


class TestIdKeTiep:
    def test_tu_danh_sach_rong(self):
        assert id_ke_tiep([]) == "ID01"

    def test_tiep_sau_ID85(self):
        assert id_ke_tiep([f"ID{i:02d}" for i in range(1, 86)]) == "ID86"

    def test_vuot_qua_ba_chu_so(self):
        assert id_ke_tiep(["ID99"]) == "ID100"

    def test_khong_tai_su_dung_id_da_thu_hoi(self):
        """ID01 đã bị bỏ nhưng tuyệt đối không được cấp lại cho người khác."""
        assert id_ke_tiep(["ID02", "ID03"]) == "ID04"


class TestDongBoSoCai:
    def test_lan_dau_cap_id_lien_tuc(self):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1"), nguoi("TRẦN THỊ B", cccd="2")], CHI_BO, [])
        assert [d.id for d in kq.dong] == ["ID01", "ID02"]
        assert kq.them_moi == ["ID01", "ID02"]

    def test_chay_lai_giu_nguyen_toan_bo_id(self):
        ds = [nguoi("NGUYỄN VĂN A", cccd="1"), nguoi("TRẦN THỊ B", cccd="2")]
        lan1 = dong_bo(ds, CHI_BO, [])
        lan2 = dong_bo(ds, CHI_BO, lan1.dong)
        assert [d.id for d in lan2.dong] == [d.id for d in lan1.dong]
        assert lan2.them_moi == []

    def test_dao_thu_tu_dong_van_giu_dung_id(self):
        """Bản xuất mới có thể sắp khác thứ tự — ID phải bám theo người, không theo dòng."""
        a, b = nguoi("NGUYỄN VĂN A", cccd="1"), nguoi("TRẦN THỊ B", cccd="2")
        lan1 = dong_bo([a, b], CHI_BO, [])
        lan2 = dong_bo([b, a], CHI_BO, lan1.dong)
        theo_ten = {d.name: d.id for d in lan2.dong}
        assert theo_ten["NGUYỄN VĂN A"] == "ID01"
        assert theo_ten["TRẦN THỊ B"] == "ID02"

    def test_them_nguoi_moi_cap_id_ke_tiep(self):
        ds = [nguoi("NGUYỄN VĂN A", cccd="1")]
        lan1 = dong_bo(ds, CHI_BO, [])
        lan2 = dong_bo(ds + [nguoi("LÊ VĂN C", cccd="3")], CHI_BO, lan1.dong)
        assert lan2.them_moi == ["ID02"]
        assert len(lan2.dong) == 2

    def test_nguoi_roi_danh_sach_chi_danh_dau_khong_xoa(self):
        ds = [nguoi("NGUYỄN VĂN A", cccd="1"), nguoi("TRẦN THỊ B", cccd="2")]
        lan1 = dong_bo(ds, CHI_BO, [])
        lan2 = dong_bo(ds[:1], CHI_BO, lan1.dong)
        assert len(lan2.dong) == 2  # KHÔNG mất dòng nào
        assert lan2.roi_danh_sach == ["ID02"]
        con_lai = next(d for d in lan2.dong if d.id == "ID02")
        assert con_lai.trang_thai == mb.TRANG_THAI_ROI
        assert con_lai.folder_name  # thư mục vẫn giữ nguyên

    def test_nguoi_roi_roi_quay_lai_van_dung_id_cu(self):
        ds = [nguoi("NGUYỄN VĂN A", cccd="1"), nguoi("TRẦN THỊ B", cccd="2")]
        v1 = dong_bo(ds, CHI_BO, [])
        v2 = dong_bo(ds[:1], CHI_BO, v1.dong)
        v3 = dong_bo(ds, CHI_BO, v2.dong)
        assert {d.name: d.id for d in v3.dong}["TRẦN THỊ B"] == "ID02"
        assert v3.them_moi == []


class TestPhatHienDuLieuBan:
    def test_bat_duoc_ca_ID58(self):
        """Giá trị hỏng thật trong MAIN.xlsx: TrànThịHòngNhuận."""
        cu = DongDangVien(
            id="ID58", so_dong=58, name="TRẦN THỊ HỒNG NHUẬN",
            name_convert="TrànThịHòngNhuận",
            folder_name="099002220002_TrànThịHòngNhuận",
            unit_folder="38168053000001", cccd_id="099002220002",
        )
        kq = dong_bo(
            [nguoi("TRẦN THỊ HỒNG NHUẬN", cccd="099002220002")], CHI_BO, [cu]
        )
        truong = {s.truong: s for s in kq.du_lieu_ban}
        assert set(truong) == {"Name_convert", "Folder_name"}
        assert truong["Name_convert"].gia_tri_moi == "TranThiHongNhuan"
        assert truong["Folder_name"].gia_tri_moi == "099002220002_TranThiHongNhuan"
        assert "có dấu" in truong["Folder_name"].ly_do

    def test_du_lieu_sach_thi_khong_bao_gi(self):
        cu = DongDangVien(
            id="ID01", so_dong=1, name="NGUYỄN VĂN A", name_convert="NguyenVanA",
            folder_name="1_NguyenVanA", unit_folder="38168053000001", cccd_id="1",
        )
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [cu])
        assert kq.du_lieu_ban == []


class TestCanhBao:
    def test_thieu_ca_hai_ma_van_tao_thu_muc(self):
        """Quyết định #12 — không chặn tiến độ vì một ca thiếu dữ liệu."""
        kq = dong_bo([nguoi("LÊ THỊ THÊM")], CHI_BO, [])
        assert kq.dong[0].folder_name == "LeThiThem"
        c = [x for x in kq.canh_bao if x.muc == "canh_bao"]
        assert len(c) == 1 and "Thiếu cả số thẻ Đảng" in c[0].van_de
        assert not kq.co_loi_chan

    def test_co_cccd_la_du_khong_can_so_the_dang(self):
        """1.Dacta_fixV1: ten thu muc dung SO CCCD, khong phai so the Dang."""
        kq = dong_bo([nguoi("LƯU THỊ HÂN", cccd="099004440004")], CHI_BO, [])
        assert kq.dong[0].folder_name == "099004440004_LuuThiHan"
        assert [x for x in kq.canh_bao if x.muc == "canh_bao"] == []

    def test_thieu_cccd_thi_bao_thong_tin_va_tam_dung_so_the(self):
        kq = dong_bo([nguoi("LƯU THỊ HÂN", dcs="099004440004")], CHI_BO, [])
        c = [x for x in kq.canh_bao if x.muc == "thong_tin"]
        assert len(c) == 1 and "Chưa có số CCCD" in c[0].van_de
        assert kq.dong[0].folder_name == "099004440004_LuuThiHan"

    def test_cccd_khong_du_12_so_thi_canh_bao(self):
        kq = dong_bo([nguoi("LƯU THỊ HÂN", cccd="12345")], CHI_BO, [])
        c = [x for x in kq.canh_bao if "12 số" in x.van_de]
        assert len(c) == 1
        assert kq.dong[0].folder_name == "12345_LuuThiHan"  # van tao, khong chan

    def test_chi_bo_la_thi_bao_loi_chan(self):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1", chi_bo="Chi bộ Không Tồn Tại")], CHI_BO, [])
        assert kq.co_loi_chan
        assert "NAME_FOLDER" in kq.canh_bao[0].van_de

    def test_uu_tien_cccd_hon_so_the_dang(self):
        """Doi chieu du lieu that: 74/85 dong hai so trung khit, 0 dong lech.

        Doi sang CCCD khong lam doi ket qua nao nhung dung 1.Dacta_fixV1.
        """
        kq = dong_bo([nguoi("NGUYỄN VĂN A", dcs="111", cccd="222")], CHI_BO, [])
        assert kq.dong[0].folder_name.startswith("222_")


class TestGhiMain:
    def test_ghi_roi_doc_lai_khop_nguyen_ven(self, tmp_path):
        kq = dong_bo([nguoi("SẦM THỊ QUỲNH NHƯ", cccd="099001110001")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, kq.dong, CHI_BO)
        doc_lai = doc_main(dich)
        assert len(doc_lai) == 1
        assert doc_lai[0].id == "ID01"
        assert doc_lai[0].folder_name == "099001110001_SamThiQuynhNhu"
        assert doc_lai[0].cccd_id == "099001110001"  # số 0 đứng đầu không bị mất

    def test_du_19_cot(self, tmp_path):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, kq.dong, CHI_BO)
        from openpyxl import load_workbook

        wb = load_workbook(dich)
        try:
            assert [c.value for c in wb[mb.SHEET_DS][1]] == mb.COT_MAIN
            assert len(mb.COT_MAIN) == 19
        finally:
            wb.close()

    def test_ma_chi_bo_ghi_lai_dung_dang_co_dau_cham(self, tmp_path):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, kq.dong, CHI_BO)
        assert doc_chi_bo(dich)["Chi bộ Văn phòng"] == ("A", "38.168.053.000.001")

    def test_unit_folder_trong_so_cai_co_dau_cham(self, tmp_path):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, kq.dong, CHI_BO)
        assert doc_main(dich)[0].unit_folder == "38.168.053.000.001"

    def test_luon_giu_sheet_danh_muc_104_loai(self, tmp_path):
        """Su co 20/8/2026: ghi thieu tham so lam mat trang sheet DANH_MUC_FILE."""
        from openpyxl import load_workbook

        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, kq.dong, CHI_BO)  # KHONG truyen danh_muc
        wb = load_workbook(dich, read_only=True)
        try:
            assert mb.SHEET_DANH_MUC in wb.sheetnames
            assert wb[mb.SHEET_DANH_MUC].max_row == 105  # 1 tieu de + 104 loai
        finally:
            wb.close()

    def test_xoay_vong_ban_sao_luu(self, tmp_path):
        dich = tmp_path / "MAIN.xlsx"
        for lan in range(3):
            kq = dong_bo([nguoi(f"NGUOI {lan}", cccd=str(lan))], CHI_BO, [])
            ghi_main(dich, kq.dong, CHI_BO)
        assert (tmp_path / "MAIN.bak1.xlsx").exists()
        assert (tmp_path / "MAIN.bak2.xlsx").exists()

    def test_khong_de_lai_tep_tam(self, tmp_path):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, kq.dong, CHI_BO)
        assert list(tmp_path.glob("*.tmp.xlsx")) == []

    def test_excel_dang_khoa_thi_bao_tieng_viet(self, tmp_path, monkeypatch):
        kq = dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, [])
        dich = tmp_path / "MAIN.xlsx"

        def gia_lap_khoa(*_a, **_k):
            raise PermissionError(13, "Permission denied")

        monkeypatch.setattr(mb.os, "replace", gia_lap_khoa)
        with pytest.raises(MainDangBiKhoa) as e:
            ghi_main(dich, kq.dong, CHI_BO)
        thong_bao = str(e.value)
        assert "Đóng MAIN.xlsx trong Excel" in thong_bao
        assert "Traceback" not in thong_bao
        assert list(tmp_path.glob("*.tmp.xlsx")) == []  # dọn sạch tệp tạm

    def test_so_cai_cu_khong_bi_hong_khi_ghi_that_bai(self, tmp_path, monkeypatch):
        dich = tmp_path / "MAIN.xlsx"
        ghi_main(dich, dong_bo([nguoi("NGUYỄN VĂN A", cccd="1")], CHI_BO, []).dong, CHI_BO)
        truoc = dich.read_bytes()

        monkeypatch.setattr(
            mb.os, "replace", lambda *a, **k: (_ for _ in ()).throw(PermissionError())
        )
        with pytest.raises(MainDangBiKhoa):
            ghi_main(dich, dong_bo([nguoi("KHÁC", cccd="9")], CHI_BO, []).dong, CHI_BO)
        assert dich.read_bytes() == truoc


class TestDocNguon:
    def test_thieu_cot_thi_bao_ro_ten_cot(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "DSTTHC"
        wb.active.append(["STT", "Họ và tên"])  # thiếu 7 cột
        p = tmp_path / "xau.xlsx"
        wb.save(p)
        wb.close()
        with pytest.raises(LoiNghiepVu) as e:
            doc_ds_dangvien(p)
        assert "Số CCCD" in str(e.value)



class TestChuanHoaUnicode:
    """Hoi quy cho loi da gap that khi chay test lan dau (20/8/2026).

    Du lieu nguon cua Vien luu ten tieng Viet lan ca hai dang NFC va NFD.
    Hai dang hien thi y het nhau nhung == cho ra False. Neu khong chuan hoa
    tai cua doc, phep tra chi bo co the truot va lam ca mot chi bo bi bao
    "khong co trong bang ma".
    """

    def test_hai_dang_unicode_trong_khac_nhau_voi_python(self):
        ten = "TRẦN THỊ HỒNG NHUẬN"
        assert unicodedata.normalize("NFC", ten) != unicodedata.normalize("NFD", ten)

    def test_chuoi_doc_ra_luon_o_dang_NFC(self):
        nfd = unicodedata.normalize("NFD", "Chi bộ Văn phòng")
        assert mb._chuoi(nfd) == unicodedata.normalize("NFC", "Chi bộ Văn phòng")

    def test_tra_chi_bo_khong_truot_du_nguon_o_dang_NFD(self, tmp_path):
        """Ca hong that: DS_DANGVIEN luu NFD, NAME_FOLDER luu NFC."""
        nguoi_nfd = nguoi("NGUYỄN VĂN A", cccd="1")
        nguoi_nfd["chi_bo"] = unicodedata.normalize("NFD", "Chi bộ Văn phòng")
        chi_bo_nfc = {
            unicodedata.normalize("NFC", "Chi bộ Văn phòng"): ("A", "38168053000001")
        }
        # Chua chuan hoa thi truot:
        assert nguoi_nfd["chi_bo"] not in chi_bo_nfc
        # Sau khi qua cua doc thi khop:
        nguoi_nfd["chi_bo"] = mb._chuoi(nguoi_nfd["chi_bo"])
        kq = dong_bo([nguoi_nfd], chi_bo_nfc, [])
        assert not kq.co_loi_chan
        assert kq.dong[0].unit_folder == "38168053000001"

    def test_bo_dau_cho_cung_ket_qua_voi_ca_hai_dang(self):
        from app.core.vietnamese import pascal_case as pc

        ten = "TRẦN THỊ HỒNG NHUẬN"
        assert (
            pc(unicodedata.normalize("NFC", ten))
            == pc(unicodedata.normalize("NFD", ten))
            == "TranThiHongNhuan"
        )

@pytest.mark.skipif(not DS_THAT.exists(), reason="Không tìm thấy DS_DANGVIEN.xlsx")
class TestDuLieuThat:
    def test_doc_dung_85_dang_vien(self):
        assert len(doc_ds_dangvien(DS_THAT)) == 85

    def test_doc_dung_7_chi_bo(self):
        assert len(doc_chi_bo(MAIN_THAT)) == 7

    def test_toan_bo_85_nguoi_khop_chi_bo(self):
        kq = dong_bo(doc_ds_dangvien(DS_THAT), doc_chi_bo(MAIN_THAT), [])
        assert not kq.co_loi_chan, [c.van_de for c in kq.canh_bao if c.muc == "loi"]
        assert len(kq.dong) == 85

    def test_moi_folder_name_deu_sach_khong_dau(self):
        from app.core.vietnamese import con_dau

        kq = dong_bo(doc_ds_dangvien(DS_THAT), doc_chi_bo(MAIN_THAT), [])
        ban = [(d.id, d.folder_name) for d in kq.dong if con_dau(d.folder_name)]
        assert ban == [], f"Còn dấu trong tên thư mục: {ban}"

    def test_dung_mot_ca_thieu_ca_hai_ma(self):
        """Đúng một đảng viên thiếu cả số thẻ Đảng lẫn CCCD.

        Không ghi tên người đó vào test: tệp test sẽ đi lên kho công khai, còn
        dữ liệu thật thì không. Kiểm bằng dấu hiệu kỹ thuật là đủ chặt.
        """
        kq = dong_bo(doc_ds_dangvien(DS_THAT), doc_chi_bo(MAIN_THAT), [])
        thieu = [c for c in kq.canh_bao if "Thiếu cả số thẻ Đảng" in c.van_de]
        assert len(thieu) == 1
        assert thieu[0].ho_ten.strip()          # có ghi rõ là của ai
        assert thieu[0].muc == "canh_bao"       # cảnh báo chứ không chặn

    def test_phat_hien_dung_du_lieu_ban_trong_MAIN_that(self):
        """Đọc MAIN.xlsx thật, phải chỉ ra dòng Folder_name còn sót dấu.

        Kiểm bằng dấu hiệu kỹ thuật — giá trị cũ còn dấu tiếng Việt, giá trị mới
        sạch dấu — chứ không ghi tên người thật vào tệp test.
        """
        from app.core.vietnamese import con_dau

        kq = dong_bo(doc_ds_dangvien(DS_THAT), doc_chi_bo(MAIN_THAT), doc_main(MAIN_THAT))
        ban_ten = [s for s in kq.du_lieu_ban if s.truong == "Folder_name"]
        assert ban_ten, "MAIN.xlsx thật đang có dòng Folder_name hỏng, phải bắt được"
        assert any(con_dau(s.gia_tri_cu) for s in ban_ten)
        assert all(not con_dau(s.gia_tri_moi) for s in ban_ten)

    def test_giu_nguyen_85_id_cua_MAIN_that(self):
        cu = doc_main(MAIN_THAT)
        kq = dong_bo(doc_ds_dangvien(DS_THAT), doc_chi_bo(MAIN_THAT), cu)
        assert kq.them_moi == [], f"Không được cấp ID mới cho người đã có: {kq.them_moi}"
        assert kq.roi_danh_sach == []
        theo_cccd_cu = {d.cccd_id: d.id for d in cu if d.cccd_id}
        for d in kq.dong:
            if d.cccd_id in theo_cccd_cu:
                assert d.id == theo_cccd_cu[d.cccd_id]

    def test_do_dai_duong_dan_nam_trong_ngan_sach(self):
        kq = dong_bo(doc_ds_dangvien(DS_THAT), doc_chi_bo(MAIN_THAT), [])
        dai_nhat = max(len(d.unit_folder) + 1 + len(d.folder_name) for d in kq.dong)
        assert dai_nhat + 1 + 101 <= 160  # còn ≥100 ký tự cho đường dẫn gốc


class TestChonNhamTepExcel:
    """Chọn nhầm tệp/thư mục phải ra thông báo tiếng Việt, không được sập.

    Lỗi thật ngày 20/8/2026: người dùng đưa vào THƯ MỤC ở ô "nơi lưu sổ cái",
    openpyxl ném InvalidFileException tiếng Anh, cả yêu cầu sập thành HTTP 500,
    giao diện báo "Không liên lạc được với ứng dụng" — sai hoàn toàn hướng.
    """

    def test_thu_muc_bao_dung_la_thu_muc(self, tmp_path):
        with pytest.raises(LoiNghiepVu) as e:
            mb.mo_workbook(tmp_path)
        assert "thư mục" in str(e.value)

    def test_tep_khong_phai_excel(self, tmp_path):
        t = tmp_path / "ghi_chu.txt"
        t.write_text("xin chao", encoding="utf-8")
        with pytest.raises(LoiNghiepVu) as e:
            mb.mo_workbook(t)
        assert "không phải Excel" in str(e.value)
        assert ".txt" in str(e.value)

    def test_tep_khong_co_duoi(self, tmp_path):
        t = tmp_path / "DACTA_FIX"
        t.write_text("dac ta", encoding="utf-8")
        with pytest.raises(LoiNghiepVu) as e:
            mb.mo_workbook(t)
        assert "không có đuôi" in str(e.value)

    def test_xls_doi_cu_duoc_mach_nuoc_cach_chuyen(self, tmp_path):
        t = tmp_path / "so_cu.xls"
        t.write_bytes(b"BM")
        with pytest.raises(LoiNghiepVu) as e:
            mb.mo_workbook(t)
        assert "Save As" in str(e.value)

    def test_tep_xlsx_hong(self, tmp_path):
        t = tmp_path / "hong.xlsx"
        t.write_bytes(b"day khong phai tep zip")
        with pytest.raises(LoiNghiepVu) as e:
            mb.mo_workbook(t)
        assert "hỏng" in str(e.value)

    def test_khong_ton_tai(self, tmp_path):
        with pytest.raises(LoiNghiepVu) as e:
            mb.mo_workbook(tmp_path / "khong_co.xlsx")
        assert "Không tìm thấy tệp" in str(e.value)

    def test_doc_main_thu_muc_thi_bao_loi_con_thieu_tep_thi_tra_rong(self, tmp_path):
        with pytest.raises(LoiNghiepVu):
            doc_main(tmp_path)
        assert doc_main(tmp_path / "chua_co.xlsx") == []


class TestChuanHoaDuongDanMain:
    """Ô "nơi lưu sổ cái" nhận thư mục, nhưng chặn mọi đuôi khác .xlsx."""

    def test_thu_muc_thi_tu_them_ten_tep(self, tmp_path):
        assert mb.chuan_hoa_duong_dan_main(tmp_path) == tmp_path / "MAIN.xlsx"

    def test_thu_muc_co_gach_cuoi(self, tmp_path):
        assert mb.chuan_hoa_duong_dan_main(str(tmp_path) + os.sep) == tmp_path / "MAIN.xlsx"

    def test_thu_muc_chua_ton_tai_cung_duoc(self, tmp_path):
        chua_co = tmp_path / "kho" / "HSDV"
        assert mb.chuan_hoa_duong_dan_main(chua_co) == chua_co / "MAIN.xlsx"

    def test_duong_dan_xlsx_giu_nguyen(self, tmp_path):
        t = tmp_path / "so_cai_rieng.xlsx"
        assert mb.chuan_hoa_duong_dan_main(t) == t

    def test_bo_dau_nhay_va_khoang_trang_khi_dan_tu_windows(self, tmp_path):
        t = tmp_path / "MAIN.xlsx"
        assert mb.chuan_hoa_duong_dan_main('  "' + str(t) + '"  ') == t

    def test_tu_choi_tep_khac_de_khong_ghi_de_mat_du_lieu(self, tmp_path):
        """Bước 2 GHI ĐÈ lên đường dẫn này — trỏ nhầm là mất tệp của người dùng."""
        dac_ta = tmp_path / "1.Dacta_fixV1"
        dac_ta.write_text("dac ta goc, khong duoc ghi de", encoding="utf-8")
        with pytest.raises(LoiNghiepVu) as e:
            mb.chuan_hoa_duong_dan_main(dac_ta)
        assert "phải là tệp .xlsx" in str(e.value)
        assert dac_ta.read_text(encoding="utf-8") == "dac ta goc, khong duoc ghi de"

    def test_bo_trong(self):
        with pytest.raises(LoiNghiepVu):
            mb.chuan_hoa_duong_dan_main("")

