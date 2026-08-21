"""Trích 104 loại tài liệu thành app/data/danh_muc_file.json.

Vì sao phải trích thay vì gõ lại
--------------------------------
Tên tệp trong bảng ``N_1..N_104`` của ``LowcodeAPP.MD`` KHÔNG suy ra được từ
tên tiếng Việt bằng phép biến đổi cơ học — bảng dùng các chữ viết tắt đã được
thống nhất trong ngành:

    Ma 5:  "Quyết định kết nạp đảng viên"  ->  QD_ket_nap_dang_vien
                                               ^^ khong phai "Quyet_dinh"
    Ma 1:  "Lý lịch của người xin vào Đảng" -> Ly_lich_nguoi_xin_vao_Dang
                                               ^ bo chu "cua"

Viết tắt đang dùng: QD, GCN, NQ, CV, TB, KL, GGT, PB, HHD, CTXH, LLVT.

Script đối chiếu chéo hai nguồn độc lập và dừng ngay nếu lệch:
  1. Bảng N_ trong LowcodeAPP.MD  -> tên dùng đặt tệp
  2. Sheet DANH_MUC_FILE của MAIN.xlsx -> tên tiếng Việt đầy đủ

Chạy:  .venv\\Scripts\\python.exe scripts\\extract_danhmuc.py
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

GOC = Path(__file__).resolve().parents[2]  # ...\3.So_Hoa
NGUON_BAT = GOC / "LowcodeAPP.MD"
NGUON_XLSX = GOC / "With_APP" / "MAIN.xlsx"
DICH = Path(__file__).resolve().parents[1] / "app" / "data" / "danh_muc_file.json"

# Ranh giới 3 độ ưu tiên theo "DM tài liệu ưu tiên số hóa.docx"
MOC_UU_TIEN = ((1, 36, 1), (37, 85, 2), (86, 104, 3))

RE_DONG_N = re.compile(r'^set\s+"N_(\d+)=(.*)"\s*$', re.MULTILINE)


def doc_ten_tep() -> dict[int, str]:
    """Đọc bảng N_1..N_104 từ tệp batch."""
    noi_dung = NGUON_BAT.read_text(encoding="utf-8", errors="replace")
    ra = {int(ma): ten.strip() for ma, ten in RE_DONG_N.findall(noi_dung)}
    if not ra:
        sys.exit(f"LOI: khong doc duoc bang N_ tu {NGUON_BAT}")
    return ra


def doc_ten_tieng_viet() -> dict[int, str]:
    """Đọc sheet DANH_MUC_FILE từ MAIN.xlsx."""
    wb = load_workbook(NGUON_XLSX, read_only=True, data_only=True)
    try:
        ws = wb["DANH_MUC_FILE"]
        ra: dict[int, str] = {}
        for hang, (ma, ten, *_) in enumerate(ws.iter_rows(values_only=True), start=1):
            if hang == 1 or ma is None or ten is None:
                continue
            try:
                ra[int(ma)] = str(ten).strip()
            except (TypeError, ValueError):
                continue
        return ra
    finally:
        wb.close()


def do_uu_tien(ma: int) -> int:
    for dau, cuoi, muc in MOC_UU_TIEN:
        if dau <= ma <= cuoi:
            return muc
    raise ValueError(f"Ma {ma} nam ngoai 1-104")


def main() -> int:
    ten_tep = doc_ten_tep()
    ten_viet = doc_ten_tieng_viet()

    loi: list[str] = []
    if sorted(ten_tep) != list(range(1, 105)):
        loi.append(f"Bang N_ co {len(ten_tep)} muc, thieu/thua so voi 1-104")
    if sorted(ten_viet) != list(range(1, 105)):
        loi.append(f"DANH_MUC_FILE co {len(ten_viet)} muc, thieu/thua so voi 1-104")
    chi_o_bat = set(ten_tep) - set(ten_viet)
    chi_o_xlsx = set(ten_viet) - set(ten_tep)
    if chi_o_bat:
        loi.append(f"Ma chi co trong LowcodeAPP.MD: {sorted(chi_o_bat)}")
    if chi_o_xlsx:
        loi.append(f"Ma chi co trong DANH_MUC_FILE: {sorted(chi_o_xlsx)}")
    for ma, ten in ten_tep.items():
        if not ten.isascii():
            loi.append(f"Ma {ma}: ten tep con ky tu ngoai ASCII: {ten!r}")
        if re.search(r'[<>:"/\\|?*]', ten):
            loi.append(f"Ma {ma}: ten tep chua ky tu Windows cam: {ten!r}")

    if loi:
        print("DOI CHIEU THAT BAI:")
        for d in loi:
            print("  -", d)
        return 1

    muc = {
        str(ma): {
            "ten_day_du": ten_viet[ma],
            "ten_tep": ten_tep[ma],
            "uu_tien": do_uu_tien(ma),
        }
        for ma in sorted(ten_tep)
    }
    dem = {m: sum(1 for v in muc.values() if v["uu_tien"] == m) for m in (1, 2, 3)}

    DICH.parent.mkdir(parents=True, exist_ok=True)
    DICH.write_text(
        json.dumps(
            {
                "nguon": [str(NGUON_BAT.name), f"{NGUON_XLSX.name}!DANH_MUC_FILE"],
                "ngay_trich": date.today().isoformat(),
                "tong_so": len(muc),
                "so_luong_theo_uu_tien": dem,
                "do_dai_ten_tep_lon_nhat": max(len(v["ten_tep"]) for v in muc.values()),
                "muc": muc,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"OK - da ghi {len(muc)} muc vao {DICH}")
    print(f"     Uu tien 1 (ma 1-36)  : {dem[1]} loai")
    print(f"     Uu tien 2 (ma 37-85) : {dem[2]} loai")
    print(f"     Uu tien 3 (ma 86-104): {dem[3]} loai")
    print(f"     Ten tep dai nhat     : {max(len(v['ten_tep']) for v in muc.values())} ky tu")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
